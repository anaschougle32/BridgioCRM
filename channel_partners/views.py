from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.conf import settings
try:
    import openpyxl
except ImportError:
    openpyxl = None
import csv
import io
import json
import uuid
import tempfile
import os
import base64
from .models import ChannelPartner
from leads.models import Lead, LeadProjectAssociation
from .utils import _create_cp_column_mapper
from projects.models import Project
from bookings.models import Booking
from leads.models import Lead


@login_required
def cp_list(request):
    """List all channel partners"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_closing_manager() or 
            request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to view channel partners.')
        return redirect('dashboard')
    
    # Filter by status first (default: show all, but can filter)
    status_filter = request.GET.get('status', '')
    if status_filter:
        cps = ChannelPartner.objects.filter(status=status_filter)
    else:
        # Default: show all CPs (both active and inactive)
        cps = ChannelPartner.objects.all()
    
    # Fix: Use status field instead of is_active
    # cps = ChannelPartner.objects.filter(is_active=True)  # OLD - removed
    
    # Mandate Owner sees all CPs, but can filter by project
    if request.user.is_mandate_owner():
        # Filter by project if specified
        project_filter = request.GET.get('project', '')
        if project_filter:
            try:
                project = Project.objects.get(id=project_filter, mandate_owner=request.user)
                cps = cps.filter(linked_projects=project).distinct()
            except Project.DoesNotExist:
                cps = ChannelPartner.objects.none()
        # Otherwise show all CPs (no filter)
    elif request.user.is_site_head():
        # Site Head sees CPs linked to their projects OR CPs with leads/bookings for their projects
        from leads.models import Lead
        from bookings.models import Booking
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        cps = cps.filter(
            Q(linked_projects__in=site_head_projects) |
            Q(leads__project__in=site_head_projects) |
            Q(bookings__project__in=site_head_projects)
        ).distinct()
    
    # Search
    search = request.GET.get('search', '')
    if search:
        cps = cps.filter(
            Q(firm_name__icontains=search) |
            Q(cp_name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search) |
            Q(cp_unique_id__icontains=search)
        )
    
    # Filter by CP type
    cp_type = request.GET.get('cp_type', '')
    if cp_type:
        cps = cps.filter(cp_type=cp_type)
    
    # Annotate with stats - use LeadProjectAssociation instead of leads
    from leads.models import LeadProjectAssociation
    if request.user.is_mandate_owner():
        cps = cps.annotate(
            lead_count=Count('leads__project_associations', filter=Q(leads__project_associations__project__mandate_owner=request.user, leads__project_associations__is_archived=False), distinct=True),
            booking_count=Count('bookings', filter=Q(bookings__project__mandate_owner=request.user, bookings__is_archived=False)),
            total_revenue=Sum('bookings__payments__amount', filter=Q(bookings__project__mandate_owner=request.user))
        ).order_by('-total_revenue', '-booking_count')
    elif request.user.is_site_head():
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        cps = cps.annotate(
            lead_count=Count('leads__project_associations', filter=Q(leads__project_associations__project__in=site_head_projects, leads__project_associations__is_archived=False), distinct=True),
            booking_count=Count('bookings', filter=Q(bookings__project__in=site_head_projects, bookings__is_archived=False)),
            total_revenue=Sum('bookings__payments__amount', filter=Q(bookings__project__in=site_head_projects))
        ).order_by('-total_revenue', '-booking_count')
    else:
        cps = cps.annotate(
            lead_count=Count('leads__project_associations', filter=Q(leads__project_associations__is_archived=False), distinct=True),
            booking_count=Count('bookings', filter=Q(bookings__is_archived=False)),
            total_revenue=Sum('bookings__payments__amount')
        ).order_by('-total_revenue', '-booking_count')
    
    # Pagination
    paginator = Paginator(cps, 25)
    page = request.GET.get('page', 1)
    cps_page = paginator.get_page(page)
    
    context = {
        'cps': cps_page,
        'cp_type_choices': ChannelPartner.CP_TYPE_CHOICES,
        'search': search,
        'selected_cp_type': cp_type,
        'status_filter': status_filter,
    }
    return render(request, 'channel_partners/list.html', context)


@login_required
def cp_search(request):
    """Search Channel Partners - AJAX endpoint for autocomplete"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    # Search by firm name, CP name, phone, or CP ID
    channel_partners = ChannelPartner.objects.filter(
        status='active'
    ).filter(
        Q(firm_name__icontains=query) |
        Q(cp_name__icontains=query) |
        Q(phone__icontains=query) |
        Q(cp_unique_id__icontains=query)
    )[:20]  # Limit to 20 results
    
    results = []
    for cp in channel_partners:
        results.append({
            'id': cp.id,
            'firm_name': cp.firm_name,
            'cp_name': cp.cp_name,
            'phone': cp.get_formatted_phone(),
            'rera_id': cp.rera_id or '',
            'cp_unique_id': cp.cp_unique_id or '',
        })
    
    return JsonResponse({'results': results})


@login_required
def cp_detail(request, pk):
    """Channel Partner detail view"""
    cp = get_object_or_404(ChannelPartner, pk=pk)
    
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_closing_manager() or 
            request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to view channel partners.')
        return redirect('dashboard')
    
    # Permission check for Mandate Owner and Site Head
    # Mandate owners have all permissions - can view all CPs
    if request.user.is_mandate_owner():
        # Mandate owners can view all CPs, no restriction needed
        pass
    elif request.user.is_super_admin():
        # Super admins can view all CPs
        pass
    elif request.user.is_site_head():
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        if not cp.linked_projects.filter(id__in=site_head_projects).exists():
            messages.error(request, 'You do not have permission to view this channel partner.')
            return redirect('channel_partners:list')
    
    # Get stats - use LeadProjectAssociation instead of direct leads
    if request.user.is_mandate_owner():
        # Get associations for leads with this CP in mandate owner's projects
        association_ids = LeadProjectAssociation.objects.filter(
            lead__channel_partner=cp,
            project__mandate_owner=request.user,
            is_archived=False
        ).values_list('lead_id', flat=True).distinct()
        leads = Lead.objects.filter(id__in=association_ids, is_archived=False)
        bookings = cp.bookings.filter(project__mandate_owner=request.user, is_archived=False)
        from bookings.models import Payment
        total_revenue = Payment.objects.filter(
            booking__channel_partner=cp,
            booking__project__mandate_owner=request.user
        ).aggregate(total=Sum('amount'))['total'] or 0
        linked_projects = cp.linked_projects.filter(mandate_owner=request.user)
    elif request.user.is_site_head():
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        # Get associations for leads with this CP in site head's projects
        association_ids = LeadProjectAssociation.objects.filter(
            lead__channel_partner=cp,
            project__in=site_head_projects,
            is_archived=False
        ).values_list('lead_id', flat=True).distinct()
        leads = Lead.objects.filter(id__in=association_ids, is_archived=False)
        bookings = cp.bookings.filter(project__in=site_head_projects, is_archived=False)
        from bookings.models import Payment
        # Use simple Sum with float conversion to avoid string concatenation issues
        total_revenue_raw = Payment.objects.filter(
            booking__channel_partner=cp,
            booking__project__in=site_head_projects
        ).aggregate(total=Sum('amount'))['total']
        total_revenue = float(total_revenue_raw) if total_revenue_raw is not None else 0
        linked_projects = cp.linked_projects.filter(id__in=site_head_projects)
    elif request.user.is_sourcing_manager():
        # Sourcing Manager sees all CP data
        association_ids = LeadProjectAssociation.objects.filter(
            lead__channel_partner=cp,
            is_archived=False
        ).values_list('lead_id', flat=True).distinct()
        leads = Lead.objects.filter(id__in=association_ids, is_archived=False)
        bookings = cp.bookings.filter(is_archived=False)
        from bookings.models import Payment
        # Use simple Sum with float conversion to avoid string concatenation issues
        total_revenue_raw = Payment.objects.filter(booking__channel_partner=cp).aggregate(
            total=Sum('amount')
        )['total']
        total_revenue = float(total_revenue_raw) if total_revenue_raw is not None else 0
        linked_projects = cp.linked_projects.all()
    else:
        association_ids = LeadProjectAssociation.objects.filter(
            lead__channel_partner=cp,
            is_archived=False
        ).values_list('lead_id', flat=True).distinct()
        leads = Lead.objects.filter(id__in=association_ids, is_archived=False)
        bookings = cp.bookings.filter(is_archived=False)
        from bookings.models import Payment
        # Use simple Sum with float conversion to avoid string concatenation issues
        total_revenue_raw = Payment.objects.filter(booking__channel_partner=cp).aggregate(
            total=Sum('amount')
        )['total']
        total_revenue = float(total_revenue_raw) if total_revenue_raw is not None else 0
        linked_projects = cp.linked_projects.all()
    
    # Get project-wise stats for Sourcing Managers
    project_stats = []
    if request.user.is_sourcing_manager():
        for project in linked_projects:
            project_leads = leads.filter(project=project)
            project_bookings = bookings.filter(project=project)
            project_revenue = Payment.objects.filter(
                booking__channel_partner=cp,
                booking__project=project
            ).aggregate(total=Sum('amount'))['total'] or 0
            
            project_stats.append({
                'project': project,
                'leads_count': project_leads.count(),
                'bookings_count': project_bookings.count(),
                'revenue': project_revenue,
            })
    
    context = {
        'cp': cp,
        'leads_count': leads.count(),
        'bookings_count': bookings.count(),
        'total_revenue': total_revenue,
        'linked_projects': linked_projects,
        'project_stats': project_stats if request.user.is_sourcing_manager() else None,
        'is_sourcing_manager': request.user.is_sourcing_manager(),
    }
    return render(request, 'channel_partners/detail.html', context)


@login_required
def cp_upload_analyze(request):
    """Analyze uploaded CP file and return headers with auto-mapping"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
        
        # Store file in session temporarily
        session_id = str(uuid.uuid4())
        
        # Read headers
        file_name = uploaded_file.name.lower()
        headers = []
        
        if file_name.endswith('.csv'):
            try:
                decoded_file = uploaded_file.read().decode('utf-8')
            except UnicodeDecodeError:
                # Try with different encodings
                uploaded_file.seek(0)
                try:
                    decoded_file = uploaded_file.read().decode('latin-1')
                except:
                    decoded_file = uploaded_file.read().decode('utf-8', errors='ignore')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            headers = list(reader.fieldnames or [])
            # Filter out empty headers
            headers = [h for h in headers if h and h.strip()]
            # Store file content in session
            request.session[f'cp_upload_file_{session_id}'] = {
                'name': uploaded_file.name,
                'content': decoded_file,
                'type': 'csv'
            }
        else:
            if openpyxl is None:
                return JsonResponse({'success': False, 'error': 'openpyxl not installed'}, status=500)
            try:
                workbook = openpyxl.load_workbook(uploaded_file, read_only=True)
                worksheet = workbook.active
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                    else:
                        headers.append('')
                # Filter out empty headers
                headers = [h for h in headers if h and h.strip()]
                workbook.close()
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Error reading Excel file: {str(e)}'}, status=500)
            # Store file in session - encode bytes to base64 for JSON serialization
            uploaded_file.seek(0)
            file_content = uploaded_file.read()
            import base64
            request.session[f'cp_upload_file_{session_id}'] = {
                'name': uploaded_file.name,
                'content': base64.b64encode(file_content).decode('utf-8'),  # Encode bytes to base64 string
                'type': 'excel'
            }
        
        if not headers:
            return JsonResponse({'success': False, 'error': 'No headers found in file. Please ensure the first row contains column names.'}, status=400)
        
        request.session.modified = True
        
        # Auto-detect mapping using CP-specific mapper
        get_value, field_map = _create_cp_column_mapper(headers)
        
        # Convert field_map (index-based) to header-based mapping
        auto_mapping = {}
        for field, col_idx in field_map.items():
            if col_idx < len(headers):
                auto_mapping[headers[col_idx]] = field
        
        return JsonResponse({
            'success': True,
            'session_id': session_id,
            'headers': headers,
            'mapping': auto_mapping
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def cp_upload_preview(request):
    """Preview CP upload with custom mapping"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        session_id = request.POST.get('session_id')
        mapping_json = request.POST.get('mapping')
        
        if not session_id or not mapping_json:
            return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)
        
        # Get file from session
        file_data = request.session.get(f'cp_upload_file_{session_id}')
        if not file_data:
            return JsonResponse({'success': False, 'error': 'File not found in session'}, status=400)
        
        # Parse mapping - handle both string and already-parsed JSON
        try:
            if isinstance(mapping_json, str):
                mapping = json.loads(mapping_json)
            else:
                mapping = mapping_json
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'Invalid mapping JSON: {str(e)}'}, status=400)
        
        # Count rows and validate
        total_rows = 0
        valid_rows = 0
        errors = 0
        error_rows = []  # Store error details
        
        if file_data['type'] == 'csv':
            io_string = io.StringIO(file_data['content'])
            reader = csv.DictReader(io_string)
            # Create reverse mapping: field -> header
            field_to_header = {v: k for k, v in mapping.items()}
            
            for row_num, row in enumerate(reader, start=2):
                total_rows += 1
                # Check required fields using mapping
                name_header = field_to_header.get('name', '')
                firm_name_header = field_to_header.get('firm_name', '')
                phone_header = field_to_header.get('phone', '')
                
                name = row.get(name_header, '').strip() if name_header else ''
                firm_name = row.get(firm_name_header, '').strip() if firm_name_header else ''
                phone = row.get(phone_header, '').strip() if phone_header else ''
                
                # Clean phone
                if phone:
                    phone = phone.replace(' ', '').replace('-', '').replace('/', '')
                    if phone.startswith('+91'):
                        phone = phone[3:]
                    elif phone.startswith('91') and len(phone) > 10:
                        phone = phone[2:]
                
                if name and firm_name and phone:
                    valid_rows += 1
                else:
                    errors += 1
                    error_rows.append({
                        'row': row_num,
                        'error': 'Name, Firm Name, and Phone are required',
                        'data': dict(row)
                    })
        else:
            # Excel preview - decode base64 content
            import base64
            file_content_bytes = base64.b64decode(file_data['content'])
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(file_content_bytes)
                tmp_path = tmp.name
            
            workbook = None
            try:
                workbook = openpyxl.load_workbook(tmp_path, read_only=True)
                worksheet = workbook.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
                
                # Create reverse mapping: field -> header
                field_to_header = {v: k for k, v in mapping.items()}
                
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                    total_rows += 1
                    # Get name, firm_name, and phone using mapping
                    name_header = field_to_header.get('name', '')
                    firm_name_header = field_to_header.get('firm_name', '')
                    phone_header = field_to_header.get('phone', '')
                    
                    name = ''
                    firm_name = ''
                    phone = ''
                    if name_header in headers:
                        name_idx = headers.index(name_header)
                        if name_idx < len(row):
                            name = str(row[name_idx].value).strip() if row[name_idx].value else ''
                    if firm_name_header in headers:
                        firm_name_idx = headers.index(firm_name_header)
                        if firm_name_idx < len(row):
                            firm_name = str(row[firm_name_idx].value).strip() if row[firm_name_idx].value else ''
                    if phone_header in headers:
                        phone_idx = headers.index(phone_header)
                        if phone_idx < len(row):
                            phone_value = row[phone_idx].value
                            # Handle Excel numeric phone numbers (convert float to int, then to string)
                            if phone_value is not None:
                                if isinstance(phone_value, (int, float)):
                                    phone = str(int(phone_value))
                                else:
                                    phone = str(phone_value).strip()
                            else:
                                phone = ''
                            # Clean phone
                            if phone:
                                phone = phone.replace(' ', '').replace('-', '').replace('/', '').replace('.', '')
                                if phone.startswith('+91'):
                                    phone = phone[3:]
                                elif phone.startswith('91') and len(phone) > 10:
                                    phone = phone[2:]
                    
                    if name and firm_name and phone:
                        valid_rows += 1
                    else:
                        errors += 1
                        error_rows.append({
                            'row': row_num,
                            'error': 'Name, Firm Name, and Phone are required',
                            'data': {headers[i]: str(row[i].value) if i < len(row) and row[i].value else '' for i in range(len(headers))}
                        })
            finally:
                # Close workbook before deleting file (Windows requires this)
                if workbook:
                    workbook.close()
                # Safe delete with retry for Windows
                try:
                    import time
                    for attempt in range(3):
                        try:
                            os.unlink(tmp_path)
                            break
                        except (OSError, PermissionError) as e:
                            if attempt < 2:
                                time.sleep(0.1)  # Wait 100ms before retry
                            else:
                                # Last attempt failed, log but don't crash
                                pass
                except Exception:
                    pass  # Ignore deletion errors
        
        return JsonResponse({
            'success': True,
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'errors': errors,
            'error_rows': error_rows[:50]  # Limit to first 50 errors for preview
        })
    except json.JSONDecodeError as e:
        return JsonResponse({'success': False, 'error': f'JSON parsing error: {str(e)}'}, status=400)
    except KeyError as e:
        return JsonResponse({'success': False, 'error': f'Missing key in file data: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        return JsonResponse({
            'success': False, 
            'error': str(e),
            'traceback': error_trace if settings.DEBUG else None
        }, status=500)


@login_required
def cp_upload(request):
    """Upload Channel Partners via Excel/CSV with multi-step mapping UI"""
    # Permission check
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to upload channel partners.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            # Check if this is the final commit
            manual_mapping_json = request.POST.get('mapping')
            session_id = request.POST.get('session_id')
            
            if not session_id or not manual_mapping_json:
                messages.error(request, 'Missing required parameters.')
                return redirect('channel_partners:upload')
            
            # Get file from session
            file_data = request.session.get(f'cp_upload_file_{session_id}')
            if not file_data:
                messages.error(request, 'File not found. Please upload again.')
                return redirect('channel_partners:upload')
            
            # Parse manual mapping
            manual_mapping = json.loads(manual_mapping_json)
            
            # Process file
            cps_created = 0
            cps_updated = 0
            errors = []
            error_rows = []  # Store failed rows for CSV download
            
            if file_data['type'] == 'csv':
                decoded_file = file_data['content']
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                headers = reader.fieldnames or []
                
                # Create reverse mapping: field -> header
                field_to_header = {v: k for k, v in manual_mapping.items()}
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Get values using mapping
                        def get_row_value(field_name):
                            header = field_to_header.get(field_name, '')
                            if header:
                                return str(row.get(header, '')).strip() if row.get(header) else ''
                            return ''
                        
                        # Required fields
                        name = get_row_value('name')
                        firm_name = get_row_value('firm_name')
                        phone_raw = get_row_value('phone')
                        
                        # Clean phone - handle Excel numeric format
                        phone = ''
                        if phone_raw:
                            # Handle Excel numeric phone numbers (convert float to int, then to string)
                            if isinstance(phone_raw, (int, float)):
                                phone = str(int(phone_raw))
                            else:
                                phone = str(phone_raw).strip()
                            # Clean phone
                            phone = phone.replace(' ', '').replace('-', '').replace('/', '').replace('.', '')
                            if phone.startswith('+91'):
                                phone = phone[3:]
                            elif phone.startswith('91') and len(phone) > 10:
                                phone = phone[2:]
                        
                        if not name or not firm_name or not phone:
                            error_msg = f"Row {row_num}: Name, Firm Name, and Phone are required"
                            errors.append(error_msg)
                            error_rows.append({
                                'row': row_num,
                                'error': error_msg,
                                'data': dict(row)
                            })
                            continue
                        
                        # Optional fields
                        phone2 = get_row_value('phone2')
                        if phone2:
                            phone2 = phone2.replace(' ', '').replace('-', '').replace('/', '')
                            if phone2.startswith('+91'):
                                phone2 = phone2[3:]
                            elif phone2.startswith('91') and len(phone2) > 10:
                                phone2 = phone2[2:]
                        
                        locality = get_row_value('locality')
                        team_size_str = get_row_value('team_size')
                        team_size = int(team_size_str) if team_size_str and team_size_str.isdigit() else None
                        owner_name = get_row_value('owner_name')
                        owner_number = get_row_value('owner_number')
                        if owner_number:
                            owner_number = owner_number.replace(' ', '').replace('-', '').replace('/', '')
                            if owner_number.startswith('+91'):
                                owner_number = owner_number[3:]
                            elif owner_number.startswith('91') and len(owner_number) > 10:
                                owner_number = owner_number[2:]
                        
                        rera_id = get_row_value('rera_id')
                        status_str = get_row_value('status')
                        # Default to 'active' if status is empty or not explicitly set to inactive
                        if status_str:
                            status = 'active' if status_str.lower() in ['active', '1', 'yes', 'true'] else 'inactive'
                        else:
                            status = 'active'  # Default to active if not specified
                        
                        # Normalize phone numbers
                        from leads.utils import normalize_phone
                        phone = normalize_phone(phone)
                        if phone2:
                            phone2 = normalize_phone(phone2)
                        if owner_number:
                            owner_number = normalize_phone(owner_number)
                        
                        # Check if CP already exists by phone
                        cp, created = ChannelPartner.objects.get_or_create(
                            phone=phone,
                            defaults={
                                'cp_name': name,
                                'firm_name': firm_name,
                                'phone2': phone2,
                                'locality': locality,
                                'team_size': team_size,
                                'owner_name': owner_name,
                                'owner_number': owner_number,
                                'rera_id': rera_id,
                                'status': status,
                            }
                        )
                        
                        if not created:
                            # Update existing CP
                            cp.cp_name = name
                            cp.firm_name = firm_name
                            cp.phone2 = phone2
                            cp.locality = locality
                            cp.team_size = team_size
                            cp.owner_name = owner_name
                            cp.owner_number = owner_number
                            cp.rera_id = rera_id
                            cp.status = status
                            cp.save()
                            cps_updated += 1
                        else:
                            cps_created += 1
                    except Exception as e:
                        error_msg = f"Row {row_num}: {str(e)}"
                        errors.append(error_msg)
                        error_rows.append({
                            'row': row_num,
                            'error': error_msg,
                            'data': dict(row)
                        })
            else:
                # Process Excel - decode base64 content
                if openpyxl is None:
                    messages.error(request, 'openpyxl is not installed. Please install it: pip install openpyxl')
                    return redirect('channel_partners:upload')
                
                import base64
                file_content_bytes = base64.b64decode(file_data['content'])
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(file_content_bytes)
                    tmp_path = tmp.name
                
                workbook = None
                try:
                    workbook = openpyxl.load_workbook(tmp_path, read_only=True)
                    worksheet = workbook.active
                    headers = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
                    
                    # Create reverse mapping: field -> header
                    field_to_header = {v: k for k, v in manual_mapping.items()}
                    
                    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                        try:
                            # Get values using mapping
                            def get_row_value(field_name):
                                header = field_to_header.get(field_name, '')
                                if header and header in headers:
                                    col_idx = headers.index(header)
                                    if col_idx < len(row):
                                        value = row[col_idx].value
                                        if value is not None:
                                            # Handle Excel numeric phone numbers (convert float to int, then to string)
                                            if isinstance(value, (int, float)) and field_name == 'phone':
                                                return str(int(value))
                                            return str(value).strip()
                                return ''
                            
                            # Required fields
                            name = get_row_value('name')
                            firm_name = get_row_value('firm_name')
                            phone_raw = get_row_value('phone')
                            
                            # Clean phone - handle Excel numeric format
                            phone = ''
                            if phone_raw:
                                # Handle Excel numeric phone numbers (convert float to int, then to string)
                                if isinstance(phone_raw, (int, float)):
                                    phone = str(int(phone_raw))
                                else:
                                    phone = str(phone_raw).strip()
                                # Clean phone
                                phone = phone.replace(' ', '').replace('-', '').replace('/', '').replace('.', '')
                                if phone.startswith('+91'):
                                    phone = phone[3:]
                                elif phone.startswith('91') and len(phone) > 10:
                                    phone = phone[2:]
                            
                            if not name or not firm_name or not phone:
                                error_msg = f"Row {row_num}: Name, Firm Name, and Phone are required"
                                errors.append(error_msg)
                                error_rows.append({
                                    'row': row_num,
                                    'error': error_msg,
                                    'data': {headers[i]: str(row[i].value) if i < len(row) and row[i].value else '' for i in range(len(headers))}
                                })
                                continue
                            
                            # Optional fields
                            phone2 = get_row_value('phone2')
                            if phone2:
                                phone2 = phone2.replace(' ', '').replace('-', '').replace('/', '')
                                if phone2.startswith('+91'):
                                    phone2 = phone2[3:]
                                elif phone2.startswith('91') and len(phone2) > 10:
                                    phone2 = phone2[2:]
                            
                            locality = get_row_value('locality')
                            team_size_str = get_row_value('team_size')
                            team_size = int(team_size_str) if team_size_str and team_size_str.isdigit() else None
                            owner_name = get_row_value('owner_name')
                            owner_number = get_row_value('owner_number')
                            if owner_number:
                                owner_number = owner_number.replace(' ', '').replace('-', '').replace('/', '')
                                if owner_number.startswith('+91'):
                                    owner_number = owner_number[3:]
                                elif owner_number.startswith('91') and len(owner_number) > 10:
                                    owner_number = owner_number[2:]
                            
                            rera_id = get_row_value('rera_id')
                            status_str = get_row_value('status')
                            # Default to 'active' if status is empty or not explicitly set to inactive
                            if status_str:
                                status = 'active' if status_str.lower() in ['active', '1', 'yes', 'true'] else 'inactive'
                            else:
                                status = 'active'  # Default to active if not specified
                            
                            # Normalize phone numbers
                            from leads.utils import normalize_phone
                            phone = normalize_phone(phone)
                            if phone2:
                                phone2 = normalize_phone(phone2)
                            if owner_number:
                                owner_number = normalize_phone(owner_number)
                            
                            # Check if CP already exists by phone
                            cp, created = ChannelPartner.objects.get_or_create(
                                phone=phone,
                                defaults={
                                    'cp_name': name,
                                    'firm_name': firm_name,
                                    'phone2': phone2,
                                    'locality': locality,
                                    'team_size': team_size,
                                    'owner_name': owner_name,
                                    'owner_number': owner_number,
                                    'rera_id': rera_id,
                                    'status': status,
                                }
                            )
                            
                            if not created:
                                # Update existing CP
                                cp.cp_name = name
                                cp.firm_name = firm_name
                                cp.phone2 = phone2
                                cp.locality = locality
                                cp.team_size = team_size
                                cp.owner_name = owner_name
                                cp.owner_number = owner_number
                                cp.rera_id = rera_id
                                cp.status = status
                                cp.save()
                                cps_updated += 1
                            else:
                                cps_created += 1
                        except Exception as e:
                            error_msg = f"Row {row_num}: {str(e)}"
                            errors.append(error_msg)
                            error_rows.append({
                                'row': row_num,
                                'error': error_msg,
                                'data': {headers[i]: str(row[i].value) if i < len(row) and row[i].value else '' for i in range(len(headers))}
                            })
                finally:
                    # Close workbook before deleting file (Windows requires this)
                    if workbook:
                        workbook.close()
                    # Safe delete with retry for Windows
                    try:
                        import time
                        for attempt in range(3):
                            try:
                                os.unlink(tmp_path)
                                break
                            except (OSError, PermissionError) as e:
                                if attempt < 2:
                                    time.sleep(0.1)  # Wait 100ms before retry
                                else:
                                    # Last attempt failed, log but don't crash
                                    pass
                    except Exception:
                        pass  # Ignore deletion errors
            
            # Clean up session
            del request.session[f'cp_upload_file_{session_id}']
            
            # Store error rows in session for CSV download
            if error_rows:
                error_session_id = str(uuid.uuid4())
                request.session[f'cp_upload_errors_{error_session_id}'] = {
                    'errors': error_rows,
                    'headers': headers if 'headers' in locals() else list(error_rows[0]['data'].keys()) if error_rows else []
                }
                request.session.modified = True
            
            if cps_created > 0 or cps_updated > 0:
                success_msg = f'Successfully processed {cps_created} new CP(s) and updated {cps_updated} existing CP(s)!'
                messages.success(request, success_msg)
            
            if errors:
                error_msg = f"Errors: {'; '.join(errors[:10])}"
                if len(errors) > 10:
                    error_msg += f" and {len(errors) - 10} more..."
                if error_rows:
                    error_msg += f" <a href='/channel-partners/upload/errors/{error_session_id}/' class='underline'>Download Error CSV</a>"
                messages.warning(request, error_msg)
            
            return redirect('channel_partners:list')
            
        except Exception as e:
            messages.error(request, f'Error uploading file: {str(e)}')
    
    return render(request, 'channel_partners/upload.html')


@login_required
def cp_upload_errors_csv(request, session_id):
    """Download error rows as CSV"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to download error files.')
        return redirect('dashboard')
    
    error_data = request.session.get(f'cp_upload_errors_{session_id}')
    if not error_data:
        messages.error(request, 'Error data not found.')
        return redirect('channel_partners:list')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cp_upload_errors_{session_id}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = error_data['headers'] + ['Error']
    writer.writerow(headers)
    
    # Write error rows
    for error_row in error_data['errors']:
        row_data = [error_row['data'].get(h, '') for h in error_data['headers']]
        row_data.append(error_row['error'])
        writer.writerow(row_data)
    
    # Clean up session
    del request.session[f'cp_upload_errors_{session_id}']
    
    return response


@login_required
def cp_create(request):
    """Create new Channel Partner"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to create channel partners.')
        return redirect('channel_partners:list')
    
    if request.method == 'POST':
        try:
            # Normalize phone numbers
            from leads.utils import normalize_phone
            normalized_phone = normalize_phone(request.POST.get('phone'))
            normalized_phone2 = normalize_phone(request.POST.get('phone2', '')) if request.POST.get('phone2') else ''
            normalized_owner_number = normalize_phone(request.POST.get('owner_number', '')) if request.POST.get('owner_number') else ''
            
            cp = ChannelPartner.objects.create(
                cp_name=request.POST.get('cp_name'),
                firm_name=request.POST.get('firm_name'),
                phone=normalized_phone,
                phone2=normalized_phone2,
                locality=request.POST.get('locality', ''),
                team_size=int(request.POST.get('team_size')) if request.POST.get('team_size') else None,
                owner_name=request.POST.get('owner_name', ''),
                owner_number=normalized_owner_number,
                rera_id=request.POST.get('rera_id', ''),
                status=request.POST.get('status', 'active'),
                email=request.POST.get('email', ''),
                cp_type=request.POST.get('cp_type', 'broker'),
                working_area=request.POST.get('working_area', ''),
            )
            messages.success(request, f'Channel Partner {cp.cp_name} created successfully!')
            return redirect('channel_partners:detail', pk=cp.pk)
        except Exception as e:
            messages.error(request, f'Error creating channel partner: {str(e)}')
    
    context = {
        'cp_type_choices': ChannelPartner.CP_TYPE_CHOICES,
    }
    return render(request, 'channel_partners/create.html', context)


@login_required
def cp_edit(request, pk):
    """Edit Channel Partner"""
    cp = get_object_or_404(ChannelPartner, pk=pk)
    
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager()):
        messages.error(request, 'You do not have permission to edit channel partners.')
        return redirect('channel_partners:list')
    
    if request.method == 'POST':
        try:
            # Normalize phone numbers
            from leads.utils import normalize_phone
            normalized_phone = normalize_phone(request.POST.get('phone'))
            normalized_phone2 = normalize_phone(request.POST.get('phone2', '')) if request.POST.get('phone2') else ''
            normalized_owner_number = normalize_phone(request.POST.get('owner_number', '')) if request.POST.get('owner_number') else ''
            
            cp.cp_name = request.POST.get('cp_name')
            cp.firm_name = request.POST.get('firm_name')
            cp.phone = normalized_phone
            cp.phone2 = normalized_phone2
            cp.locality = request.POST.get('locality', '')
            cp.team_size = int(request.POST.get('team_size')) if request.POST.get('team_size') else None
            cp.owner_name = request.POST.get('owner_name', '')
            cp.owner_number = normalized_owner_number
            cp.rera_id = request.POST.get('rera_id', '')
            cp.status = request.POST.get('status', 'active')
            cp.email = request.POST.get('email', '')
            cp.cp_type = request.POST.get('cp_type', 'broker')
            cp.working_area = request.POST.get('working_area', '')
            cp.save()
            messages.success(request, f'Channel Partner {cp.cp_name} updated successfully!')
            return redirect('channel_partners:detail', pk=cp.pk)
        except Exception as e:
            messages.error(request, f'Error updating channel partner: {str(e)}')
    
    context = {
        'cp': cp,
        'cp_type_choices': ChannelPartner.CP_TYPE_CHOICES,
    }
    return render(request, 'channel_partners/edit.html', context)
