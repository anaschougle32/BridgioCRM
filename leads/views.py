from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from datetime import timedelta, datetime
from django.template.loader import render_to_string
try:
    import openpyxl
except ImportError:
    openpyxl = None
import csv
import io
from .models import Lead, OtpLog, CallLog, FollowUpReminder, DailyAssignmentQuota, GlobalConfiguration, LeadProjectAssociation
from projects.models import Project
from accounts.models import User
from .utils import (
    generate_otp, hash_otp, verify_otp as verify_otp_hash, get_sms_deep_link,
    get_phone_display, get_tel_link, get_whatsapp_link, get_whatsapp_templates
)


def get_lead_association(lead, project=None):
    """Helper function to get LeadProjectAssociation for a lead and project"""
    if project:
        return LeadProjectAssociation.objects.filter(lead=lead, project=project, is_archived=False).first()
    # If no project specified, return primary association
    return lead.project_associations.filter(is_archived=False).first()


@login_required
def lead_list(request):
    """List all leads with filtering - works with LeadProjectAssociation"""
    # Get associations for proper filtering (project-specific data)
    associations = LeadProjectAssociation.objects.filter(is_archived=False).select_related('lead', 'project', 'assigned_to')
    
    # Role-based filtering
    if request.user.is_super_admin() or request.user.is_mandate_owner() or (request.user.is_superuser and request.user.is_staff):
        pass  # No filtering for super admin and mandate owner
    elif request.user.is_telecaller():
        # Telecallers see only their assigned leads
        associations = associations.filter(assigned_to=request.user)
    elif request.user.is_closing_manager():
        # Closing Managers see their assigned leads AND pretagged leads in their projects
        user_projects = request.user.assigned_projects.all()
        associations = associations.filter(
            Q(assigned_to=request.user) | 
            Q(is_pretagged=True, project__in=user_projects)
        )
    elif request.user.is_site_head():
        # Site head sees leads ONLY for their assigned projects (strict isolation)
        associations = associations.filter(project__site_head=request.user)
    elif request.user.is_sourcing_manager():
        # Sourcing managers see leads they created or are assigned to
        associations = associations.filter(Q(created_by=request.user) | Q(assigned_to=request.user))
    
    # Search
    search = request.GET.get('search', '')
    if search:
        associations = associations.filter(
            Q(lead__name__icontains=search) |
            Q(lead__phone__icontains=search) |
            Q(lead__email__icontains=search)
        )
    
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        associations = associations.filter(status=status)
    
    # Filter by project
    project_id = request.GET.get('project', '')
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Filter by pretag status
    pretag_status = request.GET.get('pretag_status', '')
    if pretag_status:
        associations = associations.filter(pretag_status=pretag_status)
    
    # Filter by configuration
    configuration_id = request.GET.get('configuration', '')
    if configuration_id:
        if configuration_id == 'open_budget':
            # Open Budget: leads with no configuration and no budget
            associations = associations.filter(lead__configurations__isnull=True, lead__budget__isnull=True)
        else:
            associations = associations.filter(lead__configurations__id=configuration_id)
    
    # Filter by budget range
    budget_filter = request.GET.get('budget', '')
    if budget_filter:
        if budget_filter == 'no_budget':
            associations = associations.filter(lead__budget__isnull=True)
        elif budget_filter == 'under_50l':
            associations = associations.filter(lead__budget__lt=5000000)
        elif budget_filter == '50l_to_1cr':
            associations = associations.filter(lead__budget__gte=5000000, lead__budget__lt=10000000)
        elif budget_filter == '1cr_to_2cr':
            associations = associations.filter(lead__budget__gte=10000000, lead__budget__lt=20000000)
        elif budget_filter == 'over_2cr':
            associations = associations.filter(lead__budget__gte=20000000)
    
    # Filter by assigned_to
    assigned_to_id = request.GET.get('assigned_to', '')
    if assigned_to_id:
        associations = associations.filter(assigned_to_id=assigned_to_id)
    
    # Filter by channel partner
    cp_id = request.GET.get('channel_partner', '')
    if cp_id:
        associations = associations.filter(lead__channel_partner_id=cp_id)
    
    # Filter by date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        associations = associations.filter(created_at__date__gte=date_from)
    if date_to:
        associations = associations.filter(created_at__date__lte=date_to)
    
    # Get unique lead IDs from associations
    lead_ids = associations.values_list('lead_id', flat=True).distinct()
    
    # Get leads (filtered by archived status)
    leads = Lead.objects.filter(id__in=lead_ids, is_archived=False).order_by('-created_at')
    
    # Get reminders and callbacks for notification badges
    from datetime import datetime, timedelta
    now = timezone.now()
    today = now.date()
    
    # Prefetch reminders and call logs for each lead
    leads = leads.prefetch_related('reminders', 'call_logs')
    
    # Pagination
    paginator = Paginator(leads, 25)
    page = request.GET.get('page', 1)
    leads_page = paginator.get_page(page)
    
    # Get reminders and callbacks for each lead in the page
    lead_notifications = {}
    for lead in leads_page:
        # Get upcoming reminders (today and future)
        upcoming_reminders = lead.reminders.filter(
            is_completed=False,
            reminder_date__gte=now
        ).order_by('reminder_date')[:3]
        
        # Get overdue reminders
        overdue_reminders = lead.reminders.filter(
            is_completed=False,
            reminder_date__lt=now
        ).count()
        
        # Get today's callbacks (reminder_type field was removed)
        today_callbacks = lead.reminders.filter(
            is_completed=False,
            reminder_date__date=today
        ).count()
        
        # Get tel link for phone button
        tel_link = get_tel_link(lead.phone)
        
        lead_notifications[lead.id] = {
            'upcoming_reminders': list(upcoming_reminders),
            'overdue_count': overdue_reminders,
            'today_callbacks': today_callbacks,
            'tel_link': tel_link,
        }
    
    # Get global configurations for filter dropdown
    configurations = GlobalConfiguration.objects.filter(is_active=True).order_by('order', 'name')
    
    # Get assignees for filter
    assignees = User.objects.filter(
        role__in=['closing_manager', 'telecaller', 'sourcing_manager']
    ).order_by('username')
    
    # Get channel partners for filter
    from channel_partners.models import ChannelPartner
    channel_partners = ChannelPartner.objects.filter(status='active').order_by('cp_name')
    
    # Get call metrics for current user
    # Count calls based on: CallLog entries, Notes updates, Status updates
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=today_start.weekday())
    month_start = today_start.replace(day=1)
    
    # Count CallLog entries
    call_logs_today = CallLog.objects.filter(user=request.user, call_date__gte=today_start).count()
    call_logs_week = CallLog.objects.filter(user=request.user, call_date__gte=week_start).count()
    call_logs_month = CallLog.objects.filter(user=request.user, call_date__gte=month_start).count()
    call_logs_total = CallLog.objects.filter(user=request.user).count()
    
    # Count notes updates (from AuditLog)
    from accounts.models import AuditLog
    notes_updates_today = AuditLog.objects.filter(
        user=request.user, 
        action='notes_updated',
        created_at__gte=today_start
    ).count()
    notes_updates_week = AuditLog.objects.filter(
        user=request.user, 
        action='notes_updated',
        created_at__gte=week_start
    ).count()
    notes_updates_month = AuditLog.objects.filter(
        user=request.user, 
        action='notes_updated',
        created_at__gte=month_start
    ).count()
    notes_updates_total = AuditLog.objects.filter(
        user=request.user, 
        action='notes_updated'
    ).count()
    
    # Count status updates (from AuditLog)
    status_updates_today = AuditLog.objects.filter(
        user=request.user, 
        action='status_updated',
        created_at__gte=today_start
    ).count()
    status_updates_week = AuditLog.objects.filter(
        user=request.user, 
        action='status_updated',
        created_at__gte=week_start
    ).count()
    status_updates_month = AuditLog.objects.filter(
        user=request.user, 
        action='status_updated',
        created_at__gte=month_start
    ).count()
    status_updates_total = AuditLog.objects.filter(
        user=request.user, 
        action='status_updated'
    ).count()
    
    # Combine all call activities
    call_metrics = {
        'today': call_logs_today + notes_updates_today + status_updates_today,
        'this_week': call_logs_week + notes_updates_week + status_updates_week,
        'this_month': call_logs_month + notes_updates_month + status_updates_month,
        'total': call_logs_total + notes_updates_total + status_updates_total,
    }
    
    # Generate budget choices for dropdown (common budget ranges)
    budget_choices = [
        ('2000000', '₹20L'),
        ('2500000', '₹25L'),
        ('3000000', '₹30L'),
        ('3500000', '₹35L'),
        ('4000000', '₹40L'),
        ('4500000', '₹45L'),
        ('5000000', '₹50L'),
        ('5500000', '₹55L'),
        ('6000000', '₹60L'),
        ('6500000', '₹65L'),
        ('7000000', '₹70L'),
        ('7500000', '₹75L'),
        ('8000000', '₹80L'),
        ('8500000', '₹85L'),
        ('9000000', '₹90L'),
        ('9500000', '₹95L'),
        ('10000000', '₹1Cr'),
        ('12000000', '₹1.2Cr'),
        ('15000000', '₹1.5Cr'),
        ('20000000', '₹2Cr'),
        ('25000000', '₹2.5Cr'),
        ('30000000', '₹3Cr'),
    ]
    
    # Get associations for each lead to display project-specific data
    lead_associations = {}
    lead_primary_associations = {}  # Primary association for each lead (for status, etc.)
    for lead in leads_page:
        # Get all associations for this lead
        associations = lead.project_associations.filter(is_archived=False).select_related('project', 'assigned_to')
        lead_associations[lead.id] = associations
        
        # Determine primary association (first one, or filtered by project if available)
        primary_assoc = None
        if project_id:
            primary_assoc = associations.filter(project_id=project_id).first()
        if not primary_assoc:
            primary_assoc = associations.first()
        lead_primary_associations[lead.id] = primary_assoc
    
    context = {
        'leads': leads_page,
        'lead_associations': lead_associations,
        'lead_primary_associations': lead_primary_associations,
        'projects': Project.objects.filter(is_active=True),
        'status_choices': LeadProjectAssociation.LEAD_STATUS_CHOICES,
        'pretag_status_choices': LeadProjectAssociation.PRETAG_STATUS_CHOICES,
        'configurations': configurations,
        'assignees': assignees,
        'channel_partners': channel_partners,
        'search': search,
        'selected_status': status,
        'selected_project': project_id,
        'selected_pretag_status': pretag_status,
        'selected_configuration': configuration_id,
        'selected_budget': budget_filter,
        'selected_assigned_to': assigned_to_id,
        'selected_channel_partner': cp_id,
        'date_from': date_from,
        'date_to': date_to,
        'lead_notifications': lead_notifications,
        'call_metrics': call_metrics,
        'budget_choices': budget_choices,
        'now': now,
        'today': today,
    }
    return render(request, 'leads/list.html', context)


@login_required
def lead_download(request):
    """Download leads as CSV with current filters"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        messages.error(request, 'You do not have permission to download leads.')
        return redirect('leads:list')
    
    leads = Lead.objects.filter(is_archived=False)
    
    # Apply same filters as lead_list - Mandate Owner has same permissions as Super Admin
    if request.user.is_super_admin() or request.user.is_mandate_owner() or (request.user.is_superuser and request.user.is_staff):
        pass
    elif request.user.is_telecaller() or request.user.is_closing_manager():
        leads = leads.filter(assigned_to=request.user)
    elif request.user.is_site_head():
        leads = leads.filter(project__site_head=request.user)
    elif request.user.is_sourcing_manager():
        leads = leads.filter(Q(created_by=request.user) | Q(assigned_to=request.user))
    
    # Apply search filter
    search = request.GET.get('search', '')
    if search:
        leads = leads.filter(
            Q(name__icontains=search) |
            Q(phone__icontains=search) |
            Q(email__icontains=search)
        )
    
    # Apply status filter
    status = request.GET.get('status', '')
    if status:
        leads = leads.filter(status=status)
    
    # Apply project filter
    project_id = request.GET.get('project', '')
    if project_id:
        leads = leads.filter(project_id=project_id)
    
    # Apply pretag status filter
    pretag_status = request.GET.get('pretag_status', '')
    if pretag_status:
        leads = leads.filter(pretag_status=pretag_status)
    
    # Apply configuration filter
    configuration_id = request.GET.get('configuration', '')
    if configuration_id:
        leads = leads.filter(configuration_id=configuration_id)
    
    # Apply budget filter
    budget_filter = request.GET.get('budget', '')
    if budget_filter:
        if budget_filter == 'no_budget':
            leads = leads.filter(budget__isnull=True)
        elif budget_filter == 'under_50l':
            leads = leads.filter(budget__lt=5000000)
        elif budget_filter == '50l_to_1cr':
            leads = leads.filter(budget__gte=5000000, budget__lt=10000000)
        elif budget_filter == '1cr_to_2cr':
            leads = leads.filter(budget__gte=10000000, budget__lt=20000000)
        elif budget_filter == 'over_2cr':
            leads = leads.filter(budget__gte=20000000)
    
    # Apply assigned_to filter
    assigned_to_id = request.GET.get('assigned_to', '')
    if assigned_to_id:
        leads = leads.filter(assigned_to_id=assigned_to_id)
    
    # Apply channel partner filter
    cp_id = request.GET.get('channel_partner', '')
    if cp_id:
        leads = leads.filter(channel_partner_id=cp_id)
    
    # Apply date range filter
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        leads = leads.filter(created_at__date__gte=date_from)
    if date_to:
        leads = leads.filter(created_at__date__lte=date_to)
    
    # Apply pretag status filter
    pretag_status = request.GET.get('pretag_status', '')
    if pretag_status:
        leads = leads.filter(pretag_status=pretag_status)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="leads_export.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Name', 'Phone', 'Email', 'Project', 'Configuration', 'Budget', 
        'Status', 'CP ID', 'CP Name', 'CP Firm', 'CP Phone', 'Notes', 'Created At'
    ])
    
    # Write data
    for lead in leads:
        budget_display = ''
        if lead.budget:
            if lead.budget >= 10000000:
                budget_display = f"₹{lead.budget / 10000000:.2f} Cr"
            else:
                budget_display = f"₹{lead.budget / 100000:.2f} L"
        
        writer.writerow([
            lead.name,
            lead.phone,
            lead.email or '',
            lead.project.name if lead.project else '',
            lead.configuration.name if lead.configuration else '',
            budget_display,
            lead.get_status_display(),
            lead.channel_partner.cp_unique_id if lead.channel_partner else '',
            lead.channel_partner.cp_name if lead.channel_partner else '',
            lead.channel_partner.firm_name if lead.channel_partner else '',
            lead.channel_partner.phone if lead.channel_partner else '',
            lead.notes or '',
            lead.created_at.strftime('%Y-%m-%d %H:%M:%S') if lead.created_at else '',
        ])
    
    return response


@login_required
def lead_create(request):
    """Create new visit (Available to all user types) - Multi-step form WITH OTP
    
    Flow: Client Info → OTP Verification → Requirements → CP Details
    OTP is sent via WhatsApp deep link for cost savings.
    """
    if request.method == 'POST':
        # Handle OTP sending (action='send_otp')
        if request.POST.get('action') == 'send_otp':
            phone = request.POST.get('phone', '').strip()
            if not phone:
                return JsonResponse({'success': False, 'error': 'Phone number is required.'}, status=400)
            
            # Normalize phone number
            from .utils import normalize_phone
            phone = normalize_phone(phone)
            
            # Generate OTP
            otp_code = generate_otp()
            otp_hash = hash_otp(otp_code)
            
            # Store OTP in session (temporary, for verification in step 2)
            request.session['new_visit_otp'] = {
                'otp_hash': otp_hash,
                'otp_code': otp_code,  # Store temporarily for WhatsApp link generation
                'phone': phone,
                'created_at': timezone.now().isoformat(),
            }
            request.session.modified = True
            
            # Get project name from session if available (from step 3)
            project_name = None
            visit_data = request.session.get('new_visit_data', {})
            if visit_data.get('project'):
                try:
                    project = Project.objects.get(id=visit_data.get('project'), is_active=True)
                    project_name = project.name
                except Project.DoesNotExist:
                    pass
            
            # Generate WhatsApp deep link
            from .sms_adapter import send_sms
            sms_response = send_sms(phone, otp_code, project_name=project_name)
            
            whatsapp_link = sms_response.get('whatsapp_link', '')
            
            return JsonResponse({
                'success': True,
                'whatsapp_link': whatsapp_link,
                'message': 'OTP sent successfully. WhatsApp will open automatically.'
            })
        
        step = request.POST.get('step', '1')
        
        if step == '1':
            # Step 1: Client Information - Store in session
            request.session['new_visit_data'] = {
                'name': request.POST.get('name'),
                'phone': request.POST.get('phone'),
                'email': request.POST.get('email', ''),
                'age': request.POST.get('age', ''),
                'gender': request.POST.get('gender', ''),
                'locality': request.POST.get('locality', ''),
                'current_residence': request.POST.get('current_residence', ''),
                'occupation': request.POST.get('occupation', ''),
                'company_name': request.POST.get('company_name', ''),
                'designation': request.POST.get('designation', ''),
            }
            request.session.modified = True
            return JsonResponse({'success': True, 'step': 2})
        
        elif step == '2':
            # Step 2: OTP Verification
            otp_code = request.POST.get('otp', '').strip()
            if not otp_code or len(otp_code) != 6:
                return JsonResponse({'success': False, 'error': 'Please enter a valid 6-digit OTP.'}, status=400)
            
            # Get stored OTP from session
            otp_data = request.session.get('new_visit_otp', {})
            if not otp_data:
                return JsonResponse({'success': False, 'error': 'No OTP found. Please send a new OTP.'}, status=400)
            
            # Verify OTP using utility function
            stored_hash = otp_data.get('otp_hash')
            if not stored_hash:
                return JsonResponse({'success': False, 'error': 'Invalid OTP session. Please send a new OTP.'}, status=400)
            
            is_valid = verify_otp_hash(otp_code, stored_hash)
            
            if not is_valid:
                return JsonResponse({'success': False, 'error': 'Invalid OTP. Please try again.'}, status=400)
            
            # OTP verified - mark as verified in session
            request.session['new_visit_otp_verified'] = True
            request.session.modified = True
            
            return JsonResponse({'success': True, 'step': 3})  # Move to Requirements
        
        elif step == '3':
            # Step 3: Requirements
            # Check if OTP was verified
            if not request.session.get('new_visit_otp_verified'):
                return JsonResponse({'success': False, 'error': 'Please verify OTP first.'}, status=400)
            
            visit_data = request.session.get('new_visit_data', {})
            project_id = request.POST.get('project')
            if not project_id:
                return JsonResponse({'success': False, 'error': 'Project is required.'}, status=400)
            
            visit_data.update({
                'project': project_id,
                'budget': request.POST.get('budget', ''),
                'purpose': request.POST.get('purpose', ''),
                'visit_type': request.POST.get('visit_type', ''),
                'is_first_visit': request.POST.get('is_first_visit', 'true'),
                'how_did_you_hear': request.POST.get('how_did_you_hear', ''),
            })
            request.session['new_visit_data'] = visit_data
            request.session.modified = True
            return JsonResponse({'success': True, 'step': 4})
        
        elif step == '4':
            # Step 4: CP (Optional) - Create lead
            # Check if OTP was verified
            if not request.session.get('new_visit_otp_verified'):
                return JsonResponse({'success': False, 'error': 'Please verify OTP first.'}, status=400)
            try:
                visit_data = request.session.get('new_visit_data', {})
                project_id = visit_data.get('project')
                
                if not project_id:
                    return JsonResponse({'success': False, 'error': 'Project is required.'}, status=400)
                
                project = Project.objects.get(id=project_id, is_active=True)
                
                # Normalize phone numbers
                from .utils import normalize_phone
                normalized_phone = normalize_phone(visit_data.get('phone'))
                normalized_cp_phone = normalize_phone(request.POST.get('cp_phone', '')) if request.POST.get('cp_phone') else ''
                
                # Check if lead exists by phone (deduplication)
                lead, lead_created = Lead.objects.get_or_create(
                    phone=normalized_phone,
                    defaults={
                        'name': visit_data.get('name'),
                        'email': visit_data.get('email', ''),
                        'age': int(visit_data.get('age')) if visit_data.get('age') else None,
                        'gender': visit_data.get('gender', ''),
                        'locality': visit_data.get('locality', ''),
                        'current_residence': visit_data.get('current_residence', ''),
                        'occupation': visit_data.get('occupation', ''),
                        'company_name': visit_data.get('company_name', ''),
                        'designation': visit_data.get('designation', ''),
                        'budget': float(visit_data.get('budget')) if visit_data.get('budget') else None,
                        'purpose': visit_data.get('purpose', ''),
                        'visit_type': visit_data.get('visit_type', ''),
                        'is_first_visit': visit_data.get('is_first_visit', 'true') == 'true',
                        'how_did_you_hear': visit_data.get('how_did_you_hear', ''),
                        'cp_firm_name': request.POST.get('cp_firm_name', ''),
                        'cp_name': request.POST.get('cp_name', ''),
                        'cp_phone': normalized_cp_phone,
                        'cp_rera_number': request.POST.get('cp_rera_number', ''),
                        'created_by': request.user,
                    }
                )
                
                # Update lead if it already existed (update name, email, etc. if provided)
                if not lead_created:
                    # Update fields if they're provided and different
                    if visit_data.get('name') and lead.name != visit_data.get('name'):
                        lead.name = visit_data.get('name')
                    if visit_data.get('email') and lead.email != visit_data.get('email'):
                        lead.email = visit_data.get('email')
                    # Update other fields as needed
                    lead.save()
                
                # Add configurations (multiple selection)
                config_ids = request.POST.getlist('configurations')
                if config_ids:
                    configs = GlobalConfiguration.objects.filter(id__in=config_ids, is_active=True)
                    lead.configurations.set(configs)
                
                # Check if phone is already verified in ANY project association (phone verification never expires)
                phone_already_verified = lead.project_associations.filter(
                    phone_verified=True,
                    is_archived=False
                ).exists()
                
                # Create or get LeadProjectAssociation for this project
                association, assoc_created = LeadProjectAssociation.objects.get_or_create(
                    lead=lead,
                    project=project,
                    defaults={
                        'status': 'visit_completed',  # Visit is completed when created via New Visit form
                        'is_pretagged': False,
                        'phone_verified': True,  # OTP was verified in step 2, or already verified elsewhere
                        'created_by': request.user,
                    }
                )
                
                # Update association if it already existed
                if not assoc_created:
                    association.status = 'visit_completed'
                    # OTP was verified in step 2, so always mark as verified
                    association.phone_verified = True
                    association.save()
                
                # Create OTP log entry for audit trail
                otp_data = request.session.get('new_visit_otp', {})
                if otp_data:
                    OtpLog.objects.create(
                        lead=lead,
                        otp_hash=otp_data.get('otp_hash', ''),
                        expires_at=timezone.now() + timedelta(minutes=5),
                        is_verified=True,
                        verified_at=timezone.now(),
                        attempts=1,
                        gateway_response=otp_data.get('gateway_response', '{}'),
                        max_attempts=3,
                    )
                
                # Clear session data
                request.session.pop('new_visit_data', None)
                request.session.pop('new_visit_otp', None)
                request.session.pop('new_visit_otp_verified', None)
                
                # Return JSON response with proper redirect URL
                try:
                    redirect_url = reverse('leads:detail', kwargs={'pk': lead.id})
                    # Ensure redirect_url is a valid string
                    if not redirect_url or redirect_url == 'undefined' or redirect_url == 'null':
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Invalid redirect_url generated: {redirect_url} for lead.id={lead.id}")
                        redirect_url = reverse('leads:list')
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Error generating redirect URL: {str(e)}", exc_info=True)
                    redirect_url = reverse('leads:list')
                
                return JsonResponse({
                    'success': True,
                    'message': f'New visit created successfully! Lead #{lead.id} has been added.',
                    'redirect_url': redirect_url
                })
                
            except Project.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Invalid project selected.'}, status=400)
            except ValueError as e:
                return JsonResponse({'success': False, 'error': f'Invalid data: {str(e)}'}, status=400)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating new visit: {str(e)}", exc_info=True)
                return JsonResponse({'success': False, 'error': f'Error creating new visit: {str(e)}'}, status=500)
        
        # If POST but no valid step, return error
        else:
            return JsonResponse({'success': False, 'error': 'Invalid step or request.'}, status=400)
    
    # GET request - show form
    context = {
        'projects': Project.objects.filter(is_active=True),
        'global_configurations': GlobalConfiguration.objects.filter(is_active=True).order_by('order', 'name'),
    }
    return render(request, 'leads/create.html', context)


@login_required
def search_channel_partners(request):
    """Search channel partners API endpoint for autocomplete"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    from channel_partners.models import ChannelPartner
    
    # Search by firm name, CP name, phone, or CP ID
    channel_partners = ChannelPartner.objects.filter(
        status='active'
    ).filter(
        Q(firm_name__icontains=query) |
        Q(cp_name__icontains=query) |
        Q(phone__icontains=query) |
        Q(cp_unique_id__icontains=query)
    )[:20]  # Limit to 20 results
    
    results = []
    for cp in channel_partners:
        results.append({
            'id': cp.id,
            'firm_name': cp.firm_name,
            'cp_name': cp.cp_name,
            'phone': cp.get_formatted_phone(),
            'rera_id': cp.rera_id or '',
            'cp_unique_id': cp.cp_unique_id or '',
        })
    
    return JsonResponse({'results': results})


@login_required
def search_leads(request):
    """Search leads API endpoint for autocomplete"""
    query = request.GET.get('q', '').strip()
    
    if not query:
        return JsonResponse({'results': []})
    
    # Search by name, phone, or email - show ALL leads (for budget dropdown and general search)
    # This allows searching across all projects for deduplication
    leads = Lead.objects.filter(
        is_archived=False
    ).filter(
        Q(name__icontains=query) |
        Q(phone__icontains=query) |
        Q(email__icontains=query)
    )[:20]
    
    results = []
    for lead in leads:
        # Get primary project or all projects
        primary_project = lead.primary_project
        all_projects = lead.all_projects
        
        # Get primary association status
        primary_association = lead.project_associations.filter(is_archived=False).first()
        status = primary_association.status if primary_association else 'new'
        
        results.append({
            'id': lead.id,
            'name': lead.name,
            'phone': lead.phone,
            'email': lead.email or '',
            'project': primary_project.name if primary_project else (', '.join([p.name for p in all_projects[:2]]) if all_projects else ''),
            'projects': [p.name for p in all_projects],
            'status': status,
            'configurations': [c.display_name for c in lead.configurations.all()],
        })
    
    return JsonResponse({'results': results})


@login_required
def lead_pretag(request):
    """Create pretagged lead (Sourcing Manager only)"""
    if not request.user.is_sourcing_manager():
        messages.error(request, 'Only Sourcing Managers can create pretagged leads.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            # Get primary project (first selected)
            project_ids = request.POST.getlist('projects')
            if not project_ids:
                messages.error(request, 'At least one project is required.')
                return redirect('leads:pretag')
            
            primary_project = Project.objects.get(id=project_ids[0], is_active=True)
            
            # Auto-assign to closing manager assigned to this project
            from django.utils import timezone
            closing_manager = None
            # Get closing managers assigned to this project
            closing_managers = User.objects.filter(
                role='closing_manager',
                is_active=True,
                assigned_projects=primary_project
            )
            if closing_managers.exists():
                # Assign to first available closing manager (round-robin could be implemented later)
                closing_manager = closing_managers.first()
            
            # Normalize phone numbers
            from .utils import normalize_phone
            normalized_phone = normalize_phone(request.POST.get('phone'))
            normalized_cp_phone = normalize_phone(request.POST.get('cp_phone', '')) if request.POST.get('cp_phone') else ''
            
            # Check if lead exists by phone (deduplication)
            lead, lead_created = Lead.objects.get_or_create(
                phone=normalized_phone,
                defaults={
                    'name': request.POST.get('name'),
                    'email': request.POST.get('email', ''),
                    'age': int(request.POST.get('age')) if request.POST.get('age') else None,
                    'gender': request.POST.get('gender', ''),
                    'locality': request.POST.get('locality', ''),
                    'current_residence': request.POST.get('current_residence', ''),
                    'occupation': request.POST.get('occupation', ''),
                    'company_name': request.POST.get('company_name', ''),
                    'designation': request.POST.get('designation', ''),
                    'budget': float(request.POST.get('budget')) if request.POST.get('budget') else None,
                    'purpose': request.POST.get('purpose', ''),
                    'visit_type': request.POST.get('visit_type', ''),
                    'is_first_visit': request.POST.get('is_first_visit', 'true') == 'true',
                    'how_did_you_hear': request.POST.get('how_did_you_hear', ''),
                    'cp_firm_name': request.POST.get('cp_firm_name', ''),
                    'cp_name': request.POST.get('cp_name', ''),
                    'cp_phone': normalized_cp_phone,
                    'cp_rera_number': request.POST.get('cp_rera_number', ''),
                    'created_by': request.user,
                }
            )
            
            # Update lead if it already existed
            if not lead_created:
                if request.POST.get('name') and lead.name != request.POST.get('name'):
                    lead.name = request.POST.get('name')
                if request.POST.get('email') and lead.email != request.POST.get('email'):
                    lead.email = request.POST.get('email')
                lead.save()
            
            # Add configurations (multiple selection)
            config_ids = request.POST.getlist('configurations')
            if config_ids:
                configs = GlobalConfiguration.objects.filter(id__in=config_ids, is_active=True)
                lead.configurations.set(configs)
            
            # Get date and time from form
            time_frame = request.POST.get('time_frame', '')
            visit_scheduled_date_str = request.POST.get('visit_scheduled_date', '')
            visit_scheduled_time_str = request.POST.get('visit_scheduled_time', '')
            
            # Parse date and time
            visit_scheduled_date = None
            if visit_scheduled_date_str and visit_scheduled_time_str:
                try:
                    datetime_str = f"{visit_scheduled_date_str} {visit_scheduled_time_str}"
                    visit_scheduled_date = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M")
                    visit_scheduled_date = timezone.make_aware(visit_scheduled_date)
                except (ValueError, TypeError):
                    visit_scheduled_date = None
            
            # Create associations for all selected projects
            # Create ONE association per lead+project (due to unique_together constraint)
            # ALL closing managers assigned to the project will see it in their Pretagged Visits
            # The lead is assigned to the PROJECT, not to a specific closing manager
            for project_id in project_ids:
                try:
                    project = Project.objects.get(id=project_id, is_active=True)
                    
                    # Create or get ONE association for this lead+project
                    # For single project: ALL closers in that project see it
                    # For multiple projects (universal tag): ALL closers in those projects see it
                    # Each project association has independent OTP verification and visit counting
                    # assigned_to is set to None - visibility is based on project assignment only
                    association, assoc_created = LeadProjectAssociation.objects.get_or_create(
                        lead=lead,
                        project=project,
                        defaults={
                            'status': 'new',
                            'is_pretagged': True,
                            'pretag_status': 'pending_verification',
                            'phone_verified': False,
                            'assigned_to': None,  # Not assigned to specific closer - visible to all closers in project
                            'assigned_at': None,
                            'assigned_by': None,
                            'created_by': request.user,
                            'time_frame': time_frame if time_frame else None,
                            'visit_scheduled_date': visit_scheduled_date,
                        }
                    )
                    
                    # Only update if this is a NEW pretagging action
                    # Don't modify existing associations that weren't pretagged
                    if not assoc_created:
                        # Only set pretagged if it wasn't already pretagged (to avoid overwriting)
                        # This ensures we don't accidentally pretag existing non-pretagged associations
                        if not association.is_pretagged:
                            association.is_pretagged = True
                            association.pretag_status = 'pending_verification'
                            association.phone_verified = False
                            # Clear assigned_to to make it visible to all closers in project
                            association.assigned_to = None
                            association.assigned_at = None
                            association.assigned_by = None
                        # Update time frame and date if provided
                        if time_frame:
                            association.time_frame = time_frame
                        if visit_scheduled_date:
                            association.visit_scheduled_date = visit_scheduled_date
                        association.save()
                except Project.DoesNotExist:
                    continue
            
            messages.success(request, f'Pretagged lead created successfully! Lead #{lead.id} is pending OTP verification by Closing Manager.')
            return redirect('leads:detail', pk=lead.id)
            
        except Project.DoesNotExist:
            messages.error(request, 'Invalid project selected.')
        except Exception as e:
            messages.error(request, f'Error creating pretagged lead: {str(e)}')
    
    context = {
        'projects': Project.objects.filter(is_active=True),
        'global_configurations': GlobalConfiguration.objects.filter(is_active=True).order_by('order', 'name'),
    }
    return render(request, 'leads/pretag.html', context)


@login_required
def lead_detail(request, pk):
    """Lead detail view"""
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - check associations
    has_permission = False
    
    if request.user.is_telecaller():
        # Telecallers see leads assigned to them in their projects
        associations = lead.project_associations.filter(
            assigned_to=request.user,
            project__in=request.user.assigned_projects.all(),
            is_archived=False
        )
        has_permission = associations.exists()
    elif request.user.is_closing_manager():
        # Closing managers can view assigned leads, pretagged leads, scheduled visits, or visited leads
        associations = lead.project_associations.filter(
            Q(assigned_to=request.user) | 
            Q(is_pretagged=True) |
            Q(status='visit_scheduled') |
            Q(status='visit_completed') | 
            Q(phone_verified=True),
            project__in=request.user.assigned_projects.all(),
            is_archived=False
        )
        has_permission = associations.exists()
    elif request.user.is_site_head():
        # Site head sees leads in their projects
        associations = lead.project_associations.filter(
            project__site_head=request.user,
            is_archived=False
        )
        has_permission = associations.exists()
    elif request.user.is_mandate_owner():
        # Mandate owner sees leads in their projects
        associations = lead.project_associations.filter(
            project__mandate_owner=request.user,
            is_archived=False
        )
        has_permission = associations.exists()
    elif request.user.is_sourcing_manager():
        # Sourcing managers see leads they created or are assigned to
        associations = lead.project_associations.filter(
            Q(created_by=request.user) | Q(assigned_to=request.user),
            is_archived=False
        )
        has_permission = associations.exists()
    elif request.user.is_super_admin() or request.user.is_mandate_owner():
        has_permission = True
    
    # Mandate owners have all permissions
    if request.user.is_mandate_owner():
        has_permission = True
    
    if not has_permission:
        messages.error(request, 'You do not have permission to view this lead.')
        return redirect('leads:list')
    
    # Get all associations for this lead
    associations = lead.project_associations.filter(is_archived=False).select_related('project', 'assigned_to', 'assigned_by')
    
    # Get primary association (first one, or for user's project if available)
    primary_association = None
    if request.user.is_site_head():
        primary_association = associations.filter(project__site_head=request.user).first()
    elif request.user.is_closing_manager() or request.user.is_telecaller():
        primary_association = associations.filter(
            project__in=request.user.assigned_projects.all()
        ).first()
    
    if not primary_association:
        primary_association = associations.first()
    
    # Get latest OTP log for this lead
    latest_otp = lead.otp_logs.filter(is_verified=False).order_by('-created_at').first()
    
    # Get call logs
    call_logs = lead.call_logs.all()[:10]  # Last 10 calls
    
    # Get reminders
    reminders = lead.reminders.filter(is_completed=False).order_by('reminder_date')[:5]
    
    # Get WhatsApp templates
    whatsapp_templates = get_whatsapp_templates()
    
    # Get budget choices for dropdown
    budget_choices = [
        ('2000000', '₹20L'),
        ('2500000', '₹25L'),
        ('3000000', '₹30L'),
        ('3500000', '₹35L'),
        ('4000000', '₹40L'),
        ('4500000', '₹45L'),
        ('5000000', '₹50L'),
        ('5500000', '₹55L'),
        ('6000000', '₹60L'),
        ('6500000', '₹65L'),
        ('7000000', '₹70L'),
        ('7500000', '₹75L'),
        ('8000000', '₹80L'),
        ('8500000', '₹85L'),
        ('9000000', '₹90L'),
        ('9500000', '₹95L'),
        ('10000000', '₹1Cr'),
        ('12000000', '₹1.2Cr'),
        ('15000000', '₹1.5Cr'),
        ('20000000', '₹2Cr'),
        ('25000000', '₹2.5Cr'),
        ('30000000', '₹3Cr'),
    ]
    
    context = {
        'lead': lead,
        'associations': associations,
        'primary_association': primary_association,
        'latest_otp': latest_otp,
        'now': timezone.now(),
        'call_logs': call_logs,
        'reminders': reminders,
        'whatsapp_templates': whatsapp_templates,
        'phone_display': get_phone_display(lead.phone),
        'tel_link': get_tel_link(lead.phone),
        'budget_choices': budget_choices,
        'lead_source_choices': Lead.VISIT_SOURCE_CHOICES,
        'can_edit_lead_data': request.user.is_closing_manager() or request.user.is_sourcing_manager() or request.user.is_site_head() or request.user.is_mandate_owner(),
    }
    return render(request, 'leads/detail.html', context)


@login_required
def send_otp(request, pk):
    """Send OTP to lead - Only Closing Managers, Site Heads, and Super Admins"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - Closing Managers, Telecallers, Site Heads, Mandate Owners, and Super Admins can send OTP
    if not (request.user.is_closing_manager() or request.user.is_telecaller() or request.user.is_site_head() or request.user.is_super_admin() or request.user.is_mandate_owner()):
        return JsonResponse({'success': False, 'error': 'You do not have permission to send OTP.'}, status=403)
    
    # For Closing Managers and Telecallers, allow OTP for:
    # 1. Assigned leads (in their projects)
    # 2. Pretagged leads
    # 3. Scheduled visits (status='visit_scheduled')
    # 4. Visited leads (status='visit_completed' or phone_verified=True)
    # 5. For Telecallers: Any lead in their assigned projects (for queue visit feature)
    if request.user.is_closing_manager() or request.user.is_telecaller():
        # Check associations in user's projects
        user_projects = request.user.assigned_projects.all()
        associations = lead.project_associations.filter(
            project__in=user_projects,
            is_archived=False
        )
        
        # Telecallers have broader access for queue visit feature
        if request.user.is_telecaller():
            # Telecallers can send OTP to any lead (for queue visit)
            has_permission = True
        else:
            # Closing managers need specific conditions
            has_permission = associations.filter(
                Q(assigned_to=request.user) |
                Q(is_pretagged=True) |
                Q(status='visit_scheduled') |
                Q(status='visit_completed') |
                Q(phone_verified=True)
            ).exists()
        
        if not has_permission:
            return JsonResponse({'success': False, 'error': 'You can only send OTP for leads assigned to you, pretagged leads, scheduled visits, or visited leads in your projects.'}, status=403)
        
        # Get primary project for SMS link
        if associations.exists():
            primary_project = associations.first().project
        elif user_projects.exists():
            # For telecallers with new leads, use their first assigned project
            primary_project = user_projects.first()
        else:
            primary_project = lead.primary_project
    else:
        primary_project = lead.primary_project
    
    # Check if there's an active OTP that hasn't expired OR if there's a verified OTP for pretagged leads
    now = timezone.now()
    
    # Check if this is a booking context (from unit calculation page)
    referer = request.headers.get('Referer', '')
    is_booking_context = 'calculate' in referer or 'unit' in referer or request.GET.get('for_booking') == '1'
    
    # For pretagged leads, also check for verified OTPs (they don't expire)
    # BUT: For booking conversions, always generate new OTP even if previously verified
    is_pretagged = False
    if request.user.is_closing_manager():
        user_projects = request.user.assigned_projects.all()
        associations = lead.project_associations.filter(project__in=user_projects, is_archived=False)
        is_pretagged = associations.filter(is_pretagged=True).exists()
    elif request.user.is_site_head() or request.user.is_super_admin() or request.user.is_mandate_owner():
        associations = lead.project_associations.filter(is_archived=False)
        is_pretagged = associations.filter(is_pretagged=True).exists()
    
    # Get project_id from request (GET or POST)
    project_id = request.GET.get('project_id') or request.POST.get('project_id')
    
    # Check for verified OTP first (for pretagged leads that are already verified)
    # BUT: Skip this check if it's for booking conversion - always generate new OTP for bookings
    # ALSO: For pretagged leads, OTP verification is project-specific
    # If verifying for a different project, don't show previous verified OTP
    verified_otp = None
    if is_pretagged and not is_booking_context:
        # For pretagged leads, check if there's a verified OTP for THIS specific project
        # We check by looking at the association's phone_verified status for this project
        if project_id:
            project_association = associations.filter(project_id=project_id, is_archived=False).first()
            if project_association and project_association.phone_verified:
                # This project is already verified, show verified status
                verified_otp = lead.otp_logs.filter(
                    is_verified=True
                ).order_by('-verified_at').first()
        else:
            # No specific project, check if any project is verified
            if associations.filter(phone_verified=True).exists():
                verified_otp = lead.otp_logs.filter(
                    is_verified=True
                ).order_by('-verified_at').first()
    
    # Check for active unverified OTP
    active_otp = lead.otp_logs.filter(
        is_verified=False,
        expires_at__gt=now,
        attempts__lt=3
    ).order_by('-created_at').first()
    
    # If there's a verified OTP for pretagged lead (and NOT for booking), show it as verified (no expiry)
    if verified_otp and is_pretagged and not is_booking_context:
        from .utils import get_sms_deep_link, normalize_phone
        normalized_phone = normalize_phone(lead.phone)
        project_name = primary_project.name if primary_project else ''
        sms_link = get_sms_deep_link(normalized_phone, "XXXXXX", project_name=project_name)
        
        # Return HTML showing verified status (no need to verify again)
        context = {
            'lead': lead,
            'latest_otp': verified_otp,
            'now': now,
            'sms_link': sms_link,
            'is_verified': True,
        }
        html = render_to_string('leads/otp_controls.html', context, request=request)
        return HttpResponse(html)
    
    if active_otp:
        # Generate WhatsApp link for existing OTP
        from .utils import get_sms_deep_link, normalize_phone
        normalized_phone = normalize_phone(lead.phone)
        project_name = primary_project.name if primary_project else ''
        sms_link = get_sms_deep_link(normalized_phone, "XXXXXX", project_name=project_name)
        
        # Return HTML for the OTP verification form (no OTP code shown)
        context = {
            'lead': lead,
            'latest_otp': active_otp,
            'now': now,
            'sms_link': sms_link,
        }
        html = render_to_string('leads/otp_controls.html', context, request=request)
        return HttpResponse(html)
    
    # Check if this is a JSON request (from Queue Visit or other AJAX calls)
    # Look for Accept: application/json header or Content-Type: application/json
    accept_header = request.headers.get('Accept', '')
    content_type = request.headers.get('Content-Type', '')
    is_json_request = 'application/json' in accept_header or 'application/json' in content_type
    
    try:
        # Generate OTP
        otp_code = generate_otp()
        otp_hash = hash_otp(otp_code)
        
        # Create OTP log (only hash, no plaintext)
        expires_at = now + timedelta(minutes=5)
        otp_log = OtpLog.objects.create(
            lead=lead,
            otp_hash=otp_hash,
            expires_at=expires_at,
            attempts=0,
            max_attempts=3,
            gateway_response='{}',  # Initialize with empty JSON
        )
        
        # Normalize phone number before sending
        from .utils import normalize_phone
        normalized_phone = normalize_phone(lead.phone)
        
        # Send SMS via adapter (with WhatsApp fallback)
        from .sms_adapter import send_sms
        # Get project name for SMS
        project_name = primary_project.name if primary_project else (lead.primary_project.name if lead.primary_project else '')
        sms_response = send_sms(normalized_phone, otp_code, project_name=project_name)
        
        # Store gateway response
        import json
        otp_log.gateway_response = json.dumps(sms_response)
        otp_log.save()
        
        # Generate WhatsApp link for OTP
        from .utils import get_sms_deep_link
        sms_link = get_sms_deep_link(normalized_phone, otp_code, project_name=project_name)
        
        # For JSON requests (Queue Visit), return JSON response
        if is_json_request:
            return JsonResponse({
                'success': True,
                'message': 'OTP sent successfully',
                'otp_id': otp_log.id,
                'expires_at': otp_log.expires_at.isoformat(),
                'sms_status': sms_response.get('status', 'sent'),
                'whatsapp_link': sms_response.get('whatsapp_link') if sms_response.get('status') == 'fallback' else None
            })
    except Exception as e:
        # If JSON request and error occurs, return JSON error
        if is_json_request:
            return JsonResponse({
                'success': False,
                'error': f'Error sending OTP: {str(e)}'
            }, status=500)
        # Otherwise, let Django handle the error normally
        raise
    
    # For HTML requests (lead detail page), return HTML
    context = {
        'lead': lead,
        'latest_otp': otp_log,
        'now': now,
        'sms_response': sms_response,
        'sms_link': sms_link,
    }
    html = render_to_string('leads/otp_controls.html', context, request=request)
    
    # If using WhatsApp fallback, add JavaScript to open it
    if sms_response.get('status') == 'fallback' and sms_response.get('whatsapp_link'):
        whatsapp_link = sms_response['whatsapp_link']
        html += f'''
        <script>
            // Automatically open WhatsApp when OTP is generated (fallback mode)
            (function() {{
                const whatsappLink = '{whatsapp_link}';
                if (whatsappLink) {{
                    // Open WhatsApp in new tab/window
                    window.open(whatsappLink, '_blank');
                }}
            }})();
            
            // Ensure HTMX processes the new content
            if (window.htmx) {{
                htmx.process(document.getElementById('otp-controls'));
            }}
        </script>
        '''
    
    # Return HTML directly for htmx
    return HttpResponse(html)


@login_required
def verify_otp(request, pk):
    """Verify OTP - Only Closing Managers, Site Heads, and Super Admins"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - Closing Managers, Telecallers, Site Heads, Mandate Owners, and Super Admins can verify OTP
    if not (request.user.is_closing_manager() or request.user.is_telecaller() or request.user.is_site_head() or request.user.is_super_admin() or request.user.is_mandate_owner()):
        return JsonResponse({'success': False, 'error': 'You do not have permission to verify OTP.'}, status=403)
    
    # Super Admins can verify OTP for any lead - skip project checks
    if request.user.is_super_admin():
        has_permission = True
    # For Closing Managers and Telecallers, allow OTP verification for assigned leads, scheduled visits, OR visited leads
    # For Mandate Owners and Site Heads, allow OTP verification for all leads in their projects
    elif request.user.is_closing_manager() or request.user.is_telecaller():
        # Check associations in user's projects
        user_projects = request.user.assigned_projects.all()
        associations = lead.project_associations.filter(
            project__in=user_projects,
            is_archived=False
        )
        
        # Telecallers have broader access for queue visit feature
        if request.user.is_telecaller():
            # Telecallers can verify OTP for any lead (for queue visit)
            has_permission = True
        else:
            # Closing managers need specific conditions
            has_permission = associations.filter(
                Q(assigned_to=request.user) |
                Q(is_pretagged=True) |
                Q(status='visit_scheduled') |
                Q(status='visit_completed') |
                Q(phone_verified=True)
            ).exists()
        
        if not has_permission:
            return JsonResponse({'success': False, 'error': 'You can only verify OTP for leads assigned to you, pretagged leads, scheduled visits, or visited leads in your projects.'}, status=403)
    elif request.user.is_mandate_owner():
        # Mandate owners can verify OTP for all leads (they have full access)
        # Allow verification for any lead - mandate owners have full permissions
        has_permission = True
    elif request.user.is_site_head():
        # Site heads can verify OTP for all leads in their projects
        # Also allow for booking conversion (check referer for booking context)
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        referer = request.headers.get('Referer', '')
        is_booking_context = 'calculate' in referer or 'unit' in referer
        has_permission = (
            lead.project_associations.filter(
                project__in=site_head_projects,
                is_archived=False
            ).exists() or
            # Allow for booking conversion even without association
            (is_booking_context and site_head_projects.exists())
        )
        if not has_permission:
            return JsonResponse({'success': False, 'error': 'You can only verify OTP for leads in your projects.'}, status=403)
    # Note: Super admin case is handled above with has_permission = True
    
    # Check if this is a JSON request (from Queue Visit or other AJAX calls)
    accept_header = request.headers.get('Accept', '')
    content_type = request.headers.get('Content-Type', '')
    is_json_request = 'application/json' in accept_header or 'application/json' in content_type
    
    otp_code = request.POST.get('otp', '').strip()
    if not otp_code or len(otp_code) != 6:
        return JsonResponse({'success': False, 'error': 'Invalid OTP format. Please enter a 6-digit code.'}, status=400)
    
    # Get latest unverified OTP
    now = timezone.now()
    otp_log = lead.otp_logs.filter(
        is_verified=False,
        expires_at__gt=now
    ).order_by('-created_at').first()
    
    if not otp_log:
        return JsonResponse({'success': False, 'error': 'No active OTP found. Please send a new OTP.'}, status=400)
    
    # Check attempts
    if otp_log.attempts >= otp_log.max_attempts:
        return JsonResponse({
            'success': False, 
            'error': f'Maximum attempts ({otp_log.max_attempts}) exceeded. Please send a new OTP.'
        }, status=400)
    
    # Verify OTP using utility function
    is_valid = verify_otp_hash(otp_code, otp_log.otp_hash)
    
    otp_log.attempts += 1
    
    if is_valid:
        # Mark OTP as verified
        otp_log.is_verified = True
        otp_log.verified_at = now
        
        # Get project_id from request to update specific project association
        project_id = request.POST.get('project_id') or request.GET.get('project_id')
        
        # For pretagged leads, set expires_at to far future so it never expires
        # BUT: OTP verification is project-specific for pretagged leads
        # If same lead goes to another project, they need to verify OTP again
        if project_id:
            associations = lead.project_associations.filter(project_id=project_id, is_archived=False)
        else:
            if request.user.is_closing_manager():
                user_projects = request.user.assigned_projects.all()
                associations = lead.project_associations.filter(project__in=user_projects, is_archived=False)
            else:
                associations = lead.project_associations.filter(is_archived=False)
        
        is_pretagged = associations.filter(is_pretagged=True).exists()
        if is_pretagged and project_id:
            # For pretagged leads, OTP verification is project-specific
            # Set expires_at to far future (100 years from now) so it never expires for THIS project
            from datetime import timedelta
            otp_log.expires_at = now + timedelta(days=36500)  # ~100 years
        
        otp_log.save()
        
        # Update associations - mark phone as verified and update status
        # Assign to the closing manager who verified the OTP
        for association in associations:
            # For pretagged leads, phone_verified is project-specific
            # Only mark as verified for the specific project being verified
            if association.is_pretagged and project_id:
                # Only verify for the specific project
                if str(association.project_id) != str(project_id):
                    continue
            
            association.phone_verified = True
            if association.is_pretagged:
                association.pretag_status = 'verified'
            # Update status to visit_completed when OTP is verified
            if association.status != 'booked' and association.status != 'lost':
                association.status = 'visit_completed'
            
            # Handle assignment based on user role
            if request.user.is_telecaller():
                # When telecaller verifies OTP, queue for closing managers
                association.assigned_to = None  # Unassign from telecaller
                association.pretag_status = 'queued'
                association.queued_at = timezone.now()
                association.queued_by = request.user
            elif request.user.is_closing_manager():
                # When closing manager verifies OTP, assign to themselves
                association.assigned_to = request.user
            
            association.save()
        
        # For scheduled visits, update visit count for the caller who scheduled it
        # The status change to 'visit_completed' automatically counts towards visit statistics
        
        # Create audit log
        from accounts.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='otp_verified',
            model_name='Lead',
            object_id=str(lead.id),
            changes={'lead_name': lead.name, 'phone': lead.phone, 'message': 'OTP verified'},
        )
        
        # For JSON requests (Queue Visit), return JSON response
        if is_json_request:
            return JsonResponse({
                'success': True,
                'message': 'OTP verified successfully!',
                'lead_id': lead.id,
                'phone_verified': True
            })
        
        # Return updated HTML - check if this is from booking form
        lead.refresh_from_db()
        
        # Check if request is from booking form (unit calculation page) by checking hx-target header or referer
        hx_target = request.headers.get('HX-Target', '')
        referer = request.headers.get('Referer', '')
        
        if hx_target == 'otp-booking-controls' or 'calculate' in referer or 'unit' in referer:
            # Return OTP controls with success message for booking form
            context = {
                'lead': lead,
                'latest_otp': None,
                'now': now,
            }
            html = render_to_string('leads/otp_controls.html', context, request=request)
            # Add success message at the top
            success_html = '<div class="p-3 bg-green-50 rounded-lg border border-green-200 mb-3"><p class="text-green-800 font-semibold">✓ Phone verified! You can now create the booking.</p></div>'
            html = success_html + html
            return HttpResponse(html)
        else:
            # Return full OTP section for lead detail page
            # Get primary association for the template
            primary_association = lead.project_associations.filter(is_archived=False).first()
            if not primary_association and request.user.is_closing_manager():
                # Try to get association from user's projects
                user_projects = request.user.assigned_projects.all()
                primary_association = lead.project_associations.filter(
                    project__in=user_projects,
                    is_archived=False
                ).first()
            
            context = {
                'lead': lead,
                'primary_association': primary_association,
                'latest_otp': None,
                'now': now,
            }
            html = render_to_string('leads/otp_section.html', context, request=request)
            # Add script to reload page after a short delay to update all status displays
            html += '<script>setTimeout(function(){ window.location.reload(); }, 1500);</script>'
            return HttpResponse(html)
    else:
        otp_log.save()
        remaining_attempts = otp_log.max_attempts - otp_log.attempts
        
        if remaining_attempts > 0:
            return JsonResponse({
                'success': False,
                'error': f'Invalid OTP. {remaining_attempts} attempt(s) remaining.',
            }, status=400)
        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid OTP. Maximum attempts exceeded. Please send a new OTP.',
            }, status=400)


@login_required
def visits_list(request):
    """List all visited leads (status='visit_completed') - Super Admin, Mandate Owner, Site Head only"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        messages.error(request, 'You do not have permission to view visits.')
        return redirect('dashboard')
    
    # Get filter parameters
    project_id = request.GET.get('project', '')
    visit_source = request.GET.get('visit_source', '')
    search_query = request.GET.get('search', '')
    
    # Base queryset - associations for visited leads (visit_completed) OR pretagged leads
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        associations_qs = LeadProjectAssociation.objects.filter(
            Q(status='visit_completed') | Q(is_pretagged=True),
            is_archived=False
        )
    else:  # Site Head - strict isolation
        associations_qs = LeadProjectAssociation.objects.filter(
            project__site_head=request.user,
            is_archived=False
        ).filter(
            Q(status='visit_completed') | Q(is_pretagged=True)
        )
    
    # Apply filters
    if project_id:
        associations_qs = associations_qs.filter(project_id=project_id)
    
    if visit_source:
        associations_qs = associations_qs.filter(lead__visit_source=visit_source)
    
    if search_query:
        associations_qs = associations_qs.filter(
            Q(lead__name__icontains=search_query) |
            Q(lead__phone__icontains=search_query) |
            Q(lead__email__icontains=search_query)
        )
    
    # Get projects for filter dropdown
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        projects = Project.objects.filter(is_active=True).order_by('name')
    else:  # Site Head
        projects = Project.objects.filter(site_head=request.user, is_active=True).order_by('name')
    
    # Calculate metrics
    total_visits = associations_qs.count()
    visits_by_source = associations_qs.values('lead__visit_source').annotate(count=Count('id'))
    visits_by_project = associations_qs.values('project__name').annotate(count=Count('id')).order_by('-count')[:10]
    
    # Get unique lead IDs
    lead_ids = associations_qs.values_list('lead_id', flat=True).distinct()
    
    # For Site Heads: Get assignee and handler info
    # Prefetch related data for better performance
    associations_qs = associations_qs.select_related('lead', 'assigned_to', 'created_by', 'project', 'lead__channel_partner')
    
    # Pagination - paginate associations, then get leads
    paginator = Paginator(associations_qs.order_by('-created_at'), 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get leads from associations
    association_lead_ids = [assoc.lead_id for assoc in page_obj]
    leads_qs = Lead.objects.filter(id__in=association_lead_ids, is_archived=False)
    
    # Create a mapping of lead to associations for display
    lead_associations_map = {}
    for assoc in page_obj:
        if assoc.lead_id not in lead_associations_map:
            lead_associations_map[assoc.lead_id] = []
        lead_associations_map[assoc.lead_id].append(assoc)
    
    context = {
        'associations': page_obj,  # Pass associations instead of leads
        'lead_associations_map': lead_associations_map,
        'projects': projects,
        'selected_project': project_id,
        'selected_visit_source': visit_source,
        'search_query': search_query,
        'total_visits': total_visits,
        'visits_by_source': visits_by_source,
        'visits_by_project': visits_by_project,
        'is_site_head': request.user.is_site_head(),
    }
    return render(request, 'leads/visits_list.html', context)


@login_required
def upcoming_visits(request):
    """Upcoming Visits view for Closing Managers, Mandate Owners, Site Heads, and Super Admins - shows pretagged leads and scheduled visits"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head() or request.user.is_closing_manager()):
        messages.error(request, 'You do not have permission to view upcoming visits.')
        return redirect('dashboard')
    
    # Get projects based on user role
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        assigned_projects = Project.objects.filter(is_active=True)
    elif request.user.is_site_head():
        assigned_projects = Project.objects.filter(site_head=request.user, is_active=True)
    else:  # Closing Manager
        assigned_projects = request.user.assigned_projects.filter(is_active=True)
    
    # Get associations for pretagged leads and scheduled visits
    if request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head():
        # Mandate owners and site heads see all upcoming visits in their projects
        associations = LeadProjectAssociation.objects.filter(
            project__in=assigned_projects,
            is_archived=False
        ).filter(
            Q(is_pretagged=True, pretag_status='pending_verification') |
            Q(status='visit_scheduled')
        ).select_related('lead', 'project', 'assigned_to')
    else:  # Closing Manager
        # Show pretagged leads AND scheduled visits in projects assigned to this closing manager
        # Single project pretag: ALL closers in that project see it (filtered by assigned_projects)
        # Universal tag (multiple projects): ALL closers in those projects see it
        # Each project association has independent OTP verification and visit counting
        # Pretagged leads are assigned to the PROJECT, not to a specific closing manager
        # So we only check if the project is in their assigned_projects
        # Also include scheduled visits that need OTP verification
        associations = LeadProjectAssociation.objects.filter(
            project__in=assigned_projects,
            is_archived=False
        ).filter(
            Q(is_pretagged=True, pretag_status='pending_verification') |
            Q(status='visit_scheduled', phone_verified=False)
        ).select_related('lead', 'project', 'assigned_to', 'created_by')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        associations = associations.filter(
            Q(lead__name__icontains=search) |
            Q(lead__phone__icontains=search) |
            Q(project__name__icontains=search)
        )
    
    # Filter by project
    project_id = request.GET.get('project', '')
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Filter by time frame
    time_frame = request.GET.get('time_frame', '')
    if time_frame:
        associations = associations.filter(time_frame=time_frame)
    
    # Group associations by lead_id for efficient lookup
    # This ensures we use the exact same associations we filtered, avoiding re-query mismatches
    # IMPORTANT: Evaluate the queryset here to avoid lazy evaluation issues
    all_lead_associations_dict = {}
    associations_list = list(associations.select_related('lead', 'project', 'assigned_to', 'created_by'))
    for assoc in associations_list:
        if assoc.lead_id not in all_lead_associations_dict:
            all_lead_associations_dict[assoc.lead_id] = []
        all_lead_associations_dict[assoc.lead_id].append(assoc)
    
    # Get unique lead IDs from the associations we already have
    lead_ids = list(all_lead_associations_dict.keys())
    if not lead_ids:
        # No associations found, return empty result
        leads = Lead.objects.none()
        leads_page = Paginator(leads, 25).get_page(1)
        lead_associations_dict = {}
    else:
        leads = Lead.objects.filter(id__in=lead_ids, is_archived=False).order_by('-created_at')
        
        # Pagination
        paginator = Paginator(leads, 25)
        try:
            page = int(request.GET.get('page', 1))
        except (ValueError, TypeError):
            page = 1
        leads_page = paginator.get_page(page)
        
        # Build associations dict only for leads on the current page
        # Templates can iterate over lists, so we keep them as lists
        lead_associations_dict = {}
        for lead in leads_page.object_list:  # Use object_list to ensure we iterate over actual objects
            if lead.id in all_lead_associations_dict:
                lead_associations_dict[lead.id] = all_lead_associations_dict[lead.id]
    
    # Debug output (can be removed later)
    if not lead_associations_dict and associations_list:
        # This shouldn't happen, but if it does, log it
        print(f"DEBUG: Found {len(associations_list)} associations but lead_associations_dict is empty. Lead IDs: {lead_ids[:10]}")
        # Check if leads are archived
        archived_count = Lead.objects.filter(id__in=lead_ids, is_archived=True).count()
        print(f"DEBUG: {archived_count} out of {len(lead_ids)} leads are archived")
    
    # Debug: Print what we're sending to template
    leads_count = len(leads_page.object_list) if hasattr(leads_page, 'object_list') else 0
    print(f"DEBUG upcoming_visits: associations_list={len(associations_list)}, lead_ids={len(lead_ids)}, leads_page.objects={leads_count}, lead_associations_dict={len(lead_associations_dict)}")
    if leads_page and hasattr(leads_page, 'object_list') and leads_page.object_list:
        print(f"DEBUG: First lead ID: {leads_page.object_list[0].id}, in dict: {leads_page.object_list[0].id in lead_associations_dict}")
        print(f"DEBUG: First lead associations count: {len(lead_associations_dict.get(leads_page.object_list[0].id, []))}")
    
    # Ensure we always have a valid leads_page object, even if empty
    if not leads_page:
        leads = Lead.objects.none()
        paginator = Paginator(leads, 25)
        leads_page = paginator.get_page(1)
    
    # CRITICAL FIX: Use the same pattern as pretagged_leads view which works
    # The template iterates over leads_page, so we need to ensure it's always a valid Page object
    context = {
        'leads_page': leads_page,
        'lead_associations': lead_associations_dict,
        'projects': assigned_projects,
        'search': search,
        'selected_project': project_id,
        'selected_time_frame': time_frame,
        'is_mandate_owner': request.user.is_mandate_owner(),
        'is_site_head': request.user.is_site_head(),
    }
    print(f"DEBUG FINAL: Context keys: {list(context.keys())}")
    print(f"DEBUG FINAL: leads_page exists: {'leads_page' in context}")
    print(f"DEBUG FINAL: leads_page type: {type(context.get('leads_page')).__name__ if context.get('leads_page') else 'None'}")
    print(f"DEBUG FINAL: leads_page.object_list length: {len(context['leads_page'].object_list) if context.get('leads_page') and hasattr(context['leads_page'], 'object_list') else 'N/A'}")
    print(f"DEBUG FINAL: lead_associations dict size: {len(context.get('lead_associations', {}))}")
    return render(request, 'leads/upcoming_visits.html', context)


@login_required
def pretagged_leads(request):
    """Pretagged Leads view for Sourcing Managers and Telecallers - shows pretagged leads"""
    if not (request.user.is_sourcing_manager() or request.user.is_telecaller()):
        messages.error(request, 'Only Sourcing Managers and Telecallers can view pretagged leads.')
        return redirect('dashboard')
    
    # Get projects based on user role
    if request.user.is_sourcing_manager():
        assigned_projects = request.user.assigned_projects.filter(is_active=True)
    else:  # Telecaller
        assigned_projects = Project.objects.filter(is_active=True)
    
    # Get associations for pretagged leads
    if request.user.is_sourcing_manager():
        # Sourcing managers see ALL pretagged leads in their projects
        associations = LeadProjectAssociation.objects.filter(
            is_pretagged=True,
            project__in=assigned_projects,
            is_archived=False
        ).select_related('lead', 'project', 'assigned_to', 'created_by')
    else:  # Telecaller
        # Telecallers see pretagged leads assigned to them
        associations = LeadProjectAssociation.objects.filter(
            is_pretagged=True,
            assigned_to=request.user,
            is_archived=False
        ).select_related('lead', 'project', 'assigned_to', 'created_by')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        associations = associations.filter(
            Q(lead__name__icontains=search) |
            Q(lead__phone__icontains=search) |
            Q(project__name__icontains=search) |
            Q(lead__cp_name__icontains=search) |
            Q(lead__cp_firm_name__icontains=search)
        )
    
    # Filter by project
    project_id = request.GET.get('project', '')
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Filter by visit status
    visit_status = request.GET.get('visit_status', '')
    if visit_status == 'pending':
        associations = associations.filter(pretag_status='pending_verification', phone_verified=False)
    elif visit_status == 'completed':
        associations = associations.filter(Q(pretag_status='verified') | Q(status='visit_completed'))
    
    # Get unique lead IDs
    lead_ids = associations.values_list('lead_id', flat=True).distinct()
    leads = Lead.objects.filter(id__in=lead_ids, is_archived=False).order_by('-created_at')
    
    # Get projects for filter - use assigned projects
    projects = assigned_projects.order_by('name')
    
    # Pagination
    paginator = Paginator(leads, 25)
    page = request.GET.get('page', 1)
    leads_page = paginator.get_page(page)
    
    # Get associations for each lead to display project-specific data
    lead_associations_dict = {}
    for lead in leads_page:
        lead_associations_dict[lead.id] = lead.project_associations.filter(
            is_pretagged=True,
            project__in=assigned_projects,
            is_archived=False
        ).select_related('project', 'assigned_to', 'created_by')
    
    # Stats - count associations in assigned projects
    total_pretagged = LeadProjectAssociation.objects.filter(
        is_pretagged=True,
        project__in=assigned_projects,
        is_archived=False
    ).count()
    pending_visits = LeadProjectAssociation.objects.filter(
        is_pretagged=True,
        project__in=assigned_projects,
        pretag_status='pending_verification',
        phone_verified=False,
        is_archived=False
    ).count()
    completed_visits = LeadProjectAssociation.objects.filter(
        is_pretagged=True,
        project__in=assigned_projects,
        is_archived=False
    ).filter(Q(pretag_status='verified') | Q(status='visit_completed')).count()
    
    context = {
        'leads': leads_page,
        'lead_associations': lead_associations_dict,
        'projects': projects,
        'search': search,
        'selected_project': project_id,
        'selected_visit_status': visit_status,
        'total_pretagged': total_pretagged,
        'pending_visits': pending_visits,
        'completed_visits': completed_visits,
    }
    return render(request, 'leads/pretagged_leads.html', context)


@login_required
def schedule_visit(request):
    """Schedule visit for Telecallers and Closing Managers - similar to pretagging but without CP"""
    if not (request.user.is_telecaller() or request.user.is_closing_manager()):
        messages.error(request, 'Only Telecallers and Closing Managers can schedule visits.')
        return redirect('dashboard')
    
    # Get projects assigned to this user (telecaller or closing manager)
    projects = request.user.assigned_projects.filter(is_active=True).order_by('name')
    
    if request.method == 'POST':
        try:
            # Get primary project (first selected)
            project_ids = request.POST.getlist('projects')
            if not project_ids:
                messages.error(request, 'At least one project is required.')
                return redirect('leads:schedule_visit')
            
            primary_project = Project.objects.get(id=project_ids[0], is_active=True)
            
            # Auto-assign to closing manager assigned to this project
            from django.utils import timezone
            closing_manager = None
            # Get closing managers assigned to this project
            closing_managers = User.objects.filter(
                role='closing_manager',
                is_active=True,
                assigned_projects=primary_project
            )
            if closing_managers.exists():
                # Assign to first available closing manager (round-robin could be implemented later)
                closing_manager = closing_managers.first()
            
            # Normalize phone numbers - combine country code and phone if separate
            from .utils import normalize_phone
            phone_input = request.POST.get('phone')
            country_code = request.POST.get('country_code', '+91')
            if phone_input and not phone_input.startswith('+'):
                # If phone doesn't have country code, combine with country code
                phone_input = country_code + phone_input
            normalized_phone = normalize_phone(phone_input)
            
            # Check if lead exists by phone (deduplication)
            lead, lead_created = Lead.objects.get_or_create(
                phone=normalized_phone,
                defaults={
                    'name': request.POST.get('name'),
                    'email': request.POST.get('email', ''),
                    'age': int(request.POST.get('age')) if request.POST.get('age') else None,
                    'gender': request.POST.get('gender', ''),
                    'locality': request.POST.get('locality', ''),
                    'current_residence': request.POST.get('current_residence', ''),
                    'occupation': request.POST.get('occupation', ''),
                    'company_name': request.POST.get('company_name', ''),
                    'designation': request.POST.get('designation', ''),
                    'budget': float(request.POST.get('budget')) if request.POST.get('budget') else None,
                    'purpose': request.POST.get('purpose', ''),
                    'visit_type': request.POST.get('visit_type', ''),
                    'is_first_visit': request.POST.get('is_first_visit', 'true') == 'true',
                    'how_did_you_hear': request.POST.get('how_did_you_hear', ''),
                    'created_by': request.user,
                }
            )
            
            # Update lead if it already existed
            if not lead_created:
                if request.POST.get('name') and lead.name != request.POST.get('name'):
                    lead.name = request.POST.get('name')
                lead.save()
            
            # Add configurations (multiple selection)
            config_ids = request.POST.getlist('configurations')
            if config_ids:
                configs = GlobalConfiguration.objects.filter(id__in=config_ids, is_active=True)
                lead.configurations.set(configs)
            
            # Create associations for all selected projects
            for project_id in project_ids:
                try:
                    project = Project.objects.get(id=project_id, is_active=True)
                    
                    # Auto-assign to closing manager for this project
                    project_closing_manager = None
                    project_closing_managers = User.objects.filter(
                        role='closing_manager',
                        is_active=True,
                        assigned_projects=project
                    )
                    if project_closing_managers.exists():
                        project_closing_manager = project_closing_managers.first()
                    
                    # Get time_frame and visit_scheduled_date from form
                    time_frame = request.POST.get('time_frame', '')
                    visit_scheduled_date_str = request.POST.get('visit_scheduled_date', '')
                    visit_scheduled_time_str = request.POST.get('visit_scheduled_time', '')
                    
                    visit_scheduled_date = None
                    if visit_scheduled_date_str and visit_scheduled_time_str:
                        try:
                            datetime_str = f"{visit_scheduled_date_str} {visit_scheduled_time_str}"
                            visit_scheduled_date = datetime.strptime(datetime_str, '%Y-%m-%d %H:%M')
                            visit_scheduled_date = timezone.make_aware(visit_scheduled_date)
                        except ValueError:
                            pass
                    elif visit_scheduled_date_str:
                        try:
                            date_obj = datetime.strptime(visit_scheduled_date_str, '%Y-%m-%d').date()
                            visit_scheduled_date = timezone.make_aware(datetime.combine(date_obj, datetime.min.time()))
                        except ValueError:
                            pass
                    
                    # Create or get association for this project
                    association, assoc_created = LeadProjectAssociation.objects.get_or_create(
                        lead=lead,
                        project=project,
                        defaults={
                            'status': 'visit_scheduled',
                            'is_pretagged': False,
                            'phone_verified': False,
                            'assigned_to': project_closing_manager,
                            'assigned_at': timezone.now() if project_closing_manager else None,
                            'assigned_by': request.user if project_closing_manager else None,
                            'created_by': request.user,
                            'time_frame': time_frame if time_frame else None,
                            'visit_scheduled_date': visit_scheduled_date,
                        }
                    )
                    
                    # Update association if it already existed
                    if not assoc_created:
                        if association.status != 'booked' and association.status != 'lost':
                            association.status = 'visit_scheduled'
                        if project_closing_manager and not association.assigned_to:
                            association.assigned_to = project_closing_manager
                            association.assigned_at = timezone.now()
                            association.assigned_by = request.user
                        if time_frame:
                            association.time_frame = time_frame
                        if visit_scheduled_date:
                            association.visit_scheduled_date = visit_scheduled_date
                        association.save()
                except Project.DoesNotExist:
                    continue
            
            messages.success(request, f'Visit scheduled successfully! Lead assigned to {closing_manager.username if closing_manager else "Unassigned"}.')
            return redirect('leads:scheduled_visits')
            
        except Exception as e:
            messages.error(request, f'Error scheduling visit: {str(e)}')
    
    context = {
        'projects': projects,
        'global_configurations': GlobalConfiguration.objects.filter(is_active=True).order_by('order', 'name'),
    }
    return render(request, 'leads/schedule_visit.html', context)


@login_required
def scheduled_visits(request):
    """Scheduled Visits view for Telecallers and Closing Managers - shows all scheduled visits with status"""
    if not (request.user.is_telecaller() or request.user.is_closing_manager()):
        messages.error(request, 'Only Telecallers and Closing Managers can view scheduled visits.')
        return redirect('dashboard')
    
    # Get associations for scheduled visits
    # For telecallers: show visits assigned to them or created by them
    # For closing managers: show visits assigned to them or created by them
    if request.user.is_telecaller():
        associations = LeadProjectAssociation.objects.filter(
            status='visit_scheduled',
            is_archived=False
        ).filter(
            Q(assigned_to=request.user) | Q(created_by=request.user)
        ).select_related('lead', 'project')
    else:  # Closing Manager
        # Show visits in projects assigned to this closing manager
        # This includes visits scheduled by callers for the closer to see
        assigned_projects = request.user.assigned_projects.filter(is_active=True)
        associations = LeadProjectAssociation.objects.filter(
            status='visit_scheduled',
            project__in=assigned_projects,
            is_archived=False
        ).select_related('lead', 'project', 'assigned_to', 'created_by')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        associations = associations.filter(
            Q(lead__name__icontains=search) |
            Q(lead__phone__icontains=search) |
            Q(project__name__icontains=search)
        )
    
    # Filter by project
    project_id = request.GET.get('project', '')
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Filter by time frame
    time_frame = request.GET.get('time_frame', '')
    if time_frame:
        associations = associations.filter(time_frame=time_frame)
    
    # Filter by visit scheduled date
    visit_scheduled_date = request.GET.get('visit_scheduled_date', '')
    if visit_scheduled_date:
        try:
            date_obj = datetime.strptime(visit_scheduled_date, '%Y-%m-%d').date()
            associations = associations.filter(visit_scheduled_date__date=date_obj)
        except ValueError:
            pass
    
    # Filter by status
    visit_status = request.GET.get('visit_status', '')
    if visit_status == 'pending':
        associations = associations.filter(phone_verified=False, status='visit_scheduled')
    elif visit_status == 'verified':
        associations = associations.filter(phone_verified=True)
    elif visit_status == 'completed':
        associations = associations.filter(status='visit_completed')
    
    # Get projects for filter
    projects = request.user.assigned_projects.filter(is_active=True).order_by('name')
    
    # Get unique lead IDs
    lead_ids = associations.values_list('lead_id', flat=True).distinct()
    leads = Lead.objects.filter(id__in=lead_ids, is_archived=False).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(leads, 25)
    page = request.GET.get('page', 1)
    leads_page = paginator.get_page(page)
    
    # Get associations for each lead
    lead_associations_dict = {}
    stats_filter = Q(assigned_to=request.user) | Q(created_by=request.user)
    for lead in leads_page:
        lead_associations_dict[lead.id] = lead.project_associations.filter(
            status='visit_scheduled',
            is_archived=False
        ).filter(stats_filter).select_related('project', 'assigned_to')
    
    # Stats - count associations (assigned to user or created by user)
    total_scheduled = LeadProjectAssociation.objects.filter(
        status='visit_scheduled',
        is_archived=False
    ).filter(stats_filter).count()
    pending_otp = LeadProjectAssociation.objects.filter(
        status='visit_scheduled',
        phone_verified=False,
        is_archived=False
    ).filter(stats_filter).count()
    verified = LeadProjectAssociation.objects.filter(
        status='visit_scheduled',
        phone_verified=True,
        is_archived=False
    ).filter(stats_filter).count()
    completed = LeadProjectAssociation.objects.filter(
        status='visit_completed',
        is_archived=False
    ).filter(stats_filter).count()
    
    context = {
        'leads': leads_page,
        'lead_associations': lead_associations_dict,
        'projects': projects,
        'search': search,
        'selected_project': project_id,
        'selected_time_frame': time_frame,
        'selected_visit_status': visit_status,
        'selected_visit_scheduled_date': visit_scheduled_date,
        'total_scheduled': total_scheduled,
        'pending_otp': pending_otp,
        'verified': verified,
        'completed': completed,
    }
    return render(request, 'leads/scheduled_visits.html', context)


@login_required
def closing_manager_visits(request):
    """My Visits for Closing Manager - shows visits assigned to them AND visits scheduled by callers in their projects"""
    if not request.user.is_closing_manager():
        messages.error(request, 'Only Closing Managers can view their visits.')
        return redirect('dashboard')
    
    # Get projects assigned to this closing manager
    assigned_projects = request.user.assigned_projects.filter(is_active=True)
    
    # Get associations for visits in projects assigned to this closing manager
    # This includes:
    # 1. Direct visits (visit_completed or phone_verified)
    # 2. Pretagged visits (is_pretagged=True)
    # 3. Scheduled visits (status='visit_scheduled') - scheduled by callers or directly
    # This allows closers to see all visits they handle
    associations = LeadProjectAssociation.objects.filter(
        project__in=assigned_projects,
        is_archived=False
    ).filter(
        Q(status='visit_completed') | 
        Q(phone_verified=True) | 
        Q(status='visit_scheduled') |
        Q(is_pretagged=True)
    ).select_related('lead', 'project', 'assigned_to', 'created_by')
    
    # Search
    search = request.GET.get('search', '')
    if search:
        associations = associations.filter(
            Q(lead__name__icontains=search) |
            Q(lead__phone__icontains=search) |
            Q(project__name__icontains=search)
        )
    
    # Filter by project
    project_id = request.GET.get('project', '')
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Get projects for filter
    projects = assigned_projects.order_by('name')
    
    # Get unique lead IDs
    lead_ids = associations.values_list('lead_id', flat=True).distinct()
    leads = Lead.objects.filter(id__in=lead_ids, is_archived=False).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(leads, 25)
    page = request.GET.get('page', 1)
    leads_page = paginator.get_page(page)
    
    # Get associations for each lead
    lead_associations_dict = {}
    for lead in leads_page:
        lead_associations_dict[lead.id] = lead.project_associations.filter(
            project__in=assigned_projects,
            is_archived=False
        ).filter(
            Q(status='visit_completed') | 
            Q(phone_verified=True) | 
            Q(status='visit_scheduled') |
            Q(is_pretagged=True)
        ).select_related('project', 'assigned_to', 'created_by')
    
    # Stats - count associations
    total_visits = associations.count()
    # Count visits with bookings (check if lead has booking)
    visits_with_bookings = 0
    for assoc in associations:
        if assoc.lead.bookings.exists():
            visits_with_bookings += 1
    visits_without_bookings = total_visits - visits_with_bookings
    
    context = {
        'leads': leads_page,
        'lead_associations': lead_associations_dict,
        'projects': projects,
        'search': search,
        'selected_project': project_id,
        'total_visits': total_visits,
        'visits_with_bookings': visits_with_bookings,
        'visits_without_bookings': visits_without_bookings,
    }
    return render(request, 'leads/closing_manager_visits.html', context)


@login_required
def log_call(request, pk):
    """Log a call outcome for a lead"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - check associations
    has_permission = False
    if request.user.is_telecaller() or request.user.is_closing_manager():
        # Check if user is assigned to this lead in any project
        associations = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        )
        has_permission = associations.exists()
    
    if not has_permission:
        return JsonResponse({'success': False, 'error': 'You do not have permission to log calls for this lead.'}, status=403)
    
    outcome = request.POST.get('outcome')
    notes = request.POST.get('notes', '')
    next_action = request.POST.get('next_action', '')
    
    if outcome not in dict(CallLog.OUTCOME_CHOICES):
        return JsonResponse({'success': False, 'error': 'Invalid call outcome.'}, status=400)
    
    # Create call log
    from django.utils import timezone
    call_log = CallLog.objects.create(
        lead=lead,
        user=request.user,
        call_date=timezone.now(),  # Set call_date to current time
        outcome=outcome,
        notes=notes,
    )
    
    # Auto-update status: if association is "new", change to "contacted" when called
    # Update all associations assigned to this user
    user_associations = lead.project_associations.filter(
        assigned_to=request.user,
        is_archived=False
    )
    for association in user_associations:
        if association.status == 'new':
            association.status = 'contacted'
            association.save()
    
    # Create audit log for call
    from accounts.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action='call_logged',
        model_name='Lead',
        object_id=str(lead.id),
        changes={'outcome': call_log.outcome, 'outcome_display': call_log.get_outcome_display(), 'lead_name': lead.name},
    )
    
    # Handle next action
    if next_action == 'callback':
        # Create reminder for callback
        reminder_date_str = request.POST.get('reminder_date', '')
        if reminder_date_str:
            try:
                from datetime import datetime
                reminder_date = datetime.strptime(reminder_date_str, '%Y-%m-%dT%H:%M')
                FollowUpReminder.objects.create(
                    lead=lead,
                    reminder_date=reminder_date,
                    notes=f'Callback reminder: {notes}',
                    created_by=request.user,
                )
            except ValueError:
                pass
    elif next_action == 'schedule_visit':
        # Update associations assigned to this user
        user_associations = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        )
        for association in user_associations:
            if association.status != 'booked' and association.status != 'lost':
                association.status = 'visit_scheduled'
                association.save()
    elif next_action == 'not_interested':
        # Update associations assigned to this user
        user_associations = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        )
        for association in user_associations:
            association.status = 'lost'
            association.save()
    
    # Return success response
    # For HTMX requests, return HTML that will be swapped into the target
    if request.headers.get('HX-Request'):
        success_html = (
            '<div class="p-4 bg-green-50 border border-green-200 rounded-lg">'
            '<div class="flex items-center">'
            '<svg class="w-5 h-5 text-green-600 mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">'
            '<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"></path>'
            '</svg>'
            '<p class="text-green-800 font-medium">Call logged successfully!</p>'
            '</div>'
            '<p class="text-green-600 text-sm mt-2">Reloading page to update metrics...</p>'
            '</div>'
        )
        return HttpResponse(success_html)
    # For regular requests, return JSON
    return JsonResponse({'success': True, 'message': 'Call logged successfully!'})


@login_required
def create_reminder(request, pk):
    """Create a reminder for a lead"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    reminder_date_str = request.POST.get('reminder_date', '')
    notes = request.POST.get('notes', '')
    
    if not reminder_date_str:
        return JsonResponse({'success': False, 'error': 'Reminder date is required.'}, status=400)
    
    try:
        from datetime import datetime
        reminder_date = datetime.strptime(reminder_date_str, '%Y-%m-%dT%H:%M')
        
        reminder = FollowUpReminder.objects.create(
            lead=lead,
            reminder_date=reminder_date,
            notes=notes,
            created_by=request.user,
        )
        
        # Return HTML to close modal and show success
        return HttpResponse('<div class="p-3 bg-green-50 text-green-800 rounded-lg mb-4">Reminder created successfully!</div><script>setTimeout(() => { document.getElementById("reminder-modal").classList.add("hidden"); location.reload(); }, 1500);</script>')
    
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format.'}, status=400)


@login_required
def update_status(request, pk):
    """Update lead status - updates status on LeadProjectAssociation"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Get project_id from request (required for association-based status update)
    project_id = request.POST.get('project_id') or request.GET.get('project_id')
    if not project_id:
        # Try to get primary project
        primary_project = lead.primary_project
        if not primary_project:
            return JsonResponse({'success': False, 'error': 'Project is required to update status.'}, status=400)
        project_id = primary_project.id
    
    project = get_object_or_404(Project, pk=project_id, is_active=True)
    
    # Get association for this project
    association = LeadProjectAssociation.objects.filter(
        lead=lead,
        project=project,
        is_archived=False
    ).first()
    
    if not association:
        return JsonResponse({'success': False, 'error': 'Lead is not associated with this project.'}, status=404)
    
    # Permission check - use associations for assigned_to
    has_permission = False
    
    if request.user.is_telecaller() or request.user.is_closing_manager():
        # Check if user is assigned to this lead in this project OR any project
        # Also allow if association has no assigned_to (unassigned leads)
        if association.assigned_to == request.user:
            has_permission = True
        else:
            # Check if user is assigned to this lead in any project
            other_associations = lead.project_associations.filter(
                assigned_to=request.user,
                is_archived=False
            )
            if other_associations.exists():
                has_permission = True
    elif request.user.is_mandate_owner():
        # Mandate owners can update status for leads in their projects
        has_permission = project in request.user.assigned_projects.all()
    elif request.user.is_site_head():
        # Site heads can update status for leads in their projects
        has_permission = project in request.user.assigned_projects.all()
    elif request.user.is_super_admin() or request.user.is_sourcing_manager():
        # Super admins and sourcing managers can update any lead
        has_permission = True
    else:
        has_permission = False
    
    if not has_permission:
        return JsonResponse({'success': False, 'error': 'You do not have permission to update this lead.'}, status=403)
    
    new_status = request.POST.get('status', '')
    if new_status not in dict(LeadProjectAssociation.LEAD_STATUS_CHOICES):
        return JsonResponse({'success': False, 'error': 'Invalid status.'}, status=400)
    
    # Update status on association
    old_status = association.status
    association.status = new_status
    association.save()
    
    # Create audit log
    from accounts.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action='status_updated',
        model_name='LeadProjectAssociation',
        object_id=str(association.id),
        changes={'status': new_status, 'old_status': old_status, 'status_display': association.get_status_display()},
    )
    
    messages.success(request, f'Lead status updated to {association.get_status_display()}.')
    
    # Return JSON for AJAX requests, redirect for regular form submissions
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.headers.get('Content-Type') == 'application/json':
        return JsonResponse({'success': True, 'message': f'Lead status updated to {association.get_status_display()}.'})
    
    return redirect('leads:detail', pk=lead.id)


@login_required
def complete_reminder(request, pk, reminder_id):
    """Mark a reminder as completed"""
    reminder = get_object_or_404(FollowUpReminder, pk=reminder_id, lead_id=pk)
    reminder.is_completed = True
    reminder.completed_at = timezone.now()
    reminder.save()
    messages.success(request, 'Reminder marked as completed.')
    
    # If HTMX request, return JSON for dynamic update
    is_htmx = request.headers.get('HX-Request') or request.headers.get('hx-request')
    if is_htmx:
        return JsonResponse({'success': True, 'message': 'Reminder completed'})
    
    return redirect('leads:detail', pk=pk)


@login_required
def followups_list(request):
    """List all follow-up reminders with Kanban board (desktop) and list view (mobile/tablet)"""
    now = timezone.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    
    # Get all reminders based on user role
    reminders = FollowUpReminder.objects.select_related('lead', 'created_by').all()
    
    # Role-based filtering
    if request.user.is_telecaller() or request.user.is_closing_manager():
        # See reminders for leads assigned to them
        reminders = reminders.filter(
            lead__project_associations__assigned_to=request.user,
            lead__project_associations__is_archived=False
        ).distinct()
    elif request.user.is_site_head():
        # See reminders for leads in their projects
        reminders = reminders.filter(
            lead__project_associations__project__site_head=request.user,
            lead__project_associations__is_archived=False
        ).distinct()
    elif request.user.is_sourcing_manager():
        # See reminders for leads they created or are assigned to
        reminders = reminders.filter(
            Q(created_by=request.user) |
            Q(lead__project_associations__assigned_to=request.user)
        ).distinct()
    # Super admin and mandate owner see all
    
    # Filter by assigned user if provided
    assigned_to = request.GET.get('assigned_to', '')
    if assigned_to:
        reminders = reminders.filter(
            lead__project_associations__assigned_to_id=assigned_to,
            lead__project_associations__is_archived=False
        ).distinct()
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            reminders = reminders.filter(reminder_date__gte=date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            reminders = reminders.filter(reminder_date__lt=date_to_obj)
        except ValueError:
            pass
    
    # Categorize reminders
    overdue = reminders.filter(is_completed=False, reminder_date__lt=today_start).order_by('reminder_date')
    today = reminders.filter(is_completed=False, reminder_date__gte=today_start, reminder_date__lt=today_end).order_by('reminder_date')
    upcoming = reminders.filter(is_completed=False, reminder_date__gte=today_end).order_by('reminder_date')
    completed = reminders.filter(is_completed=True).order_by('-completed_at')[:50]  # Limit completed to recent 50
    
    # Get assignees for filter
    assignees = User.objects.filter(
        role__in=['closing_manager', 'telecaller', 'sourcing_manager']
    ).order_by('username')
    
    context = {
        'overdue': overdue,
        'today': today,
        'upcoming': upcoming,
        'completed': completed,
        'assignees': assignees,
        'selected_assigned_to': assigned_to,
        'date_from': date_from,
        'date_to': date_to,
        'now': now,
    }
    return render(request, 'leads/followups.html', context)


@login_required
def update_notes(request, pk):
    """Update lead notes - append new notes with timestamp"""
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - use associations for assigned_to
    has_permission = False
    
    if request.user.is_telecaller() or request.user.is_closing_manager():
        # Check if user is assigned to this lead in any project
        has_permission = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        ).exists()
    elif request.user.is_mandate_owner() or request.user.is_site_head():
        # Mandate owners and site heads can update notes for leads in their projects
        user_projects = request.user.assigned_projects.all()
        has_permission = lead.project_associations.filter(
            project__in=user_projects,
            is_archived=False
        ).exists()
    elif request.user.is_super_admin() or request.user.is_sourcing_manager():
        # Super admins and sourcing managers can update any lead
        has_permission = True
    else:
        has_permission = False
    
    if not has_permission:
        # Return error for HTMX
        is_htmx = request.headers.get('HX-Request') or request.headers.get('hx-request')
        if is_htmx:
            error_html = (
                '<div class="p-4 bg-red-50 border border-red-200 rounded-lg">'
                '<p class="text-red-800 font-medium">Error: You do not have permission to update notes for this lead.</p>'
                '</div>'
            )
            return HttpResponse(error_html, status=403)
        return JsonResponse({'success': False, 'error': 'You do not have permission to update notes for this lead.'}, status=403)
    
    if request.method == 'GET':
        # Don't return notes - we're adding new notes, not editing existing ones
        return JsonResponse({'success': True, 'notes': ''})
    
    if request.method == 'POST':
        new_note = request.POST.get('notes', '').strip()
        if not new_note:
            # Return error for HTMX
            is_htmx = request.headers.get('HX-Request') or request.headers.get('hx-request')
            if is_htmx:
                error_html = (
                    '<div class="p-4 bg-red-50 border border-red-200 rounded-lg">'
                    '<p class="text-red-800 font-medium">Error: Note cannot be empty.</p>'
                    '</div>'
                )
                return HttpResponse(error_html, status=400)
            return JsonResponse({'success': False, 'error': 'Note cannot be empty.'}, status=400)
        
        # Append new note with timestamp
        from django.utils import timezone
        timestamp = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        user_name = request.user.get_full_name() or request.user.username
        
        if lead.notes:
            # Append to existing notes
            lead.notes = f"{lead.notes}\n\n--- {timestamp} ({user_name}) ---\n{new_note}"
        else:
            # First note
            lead.notes = f"--- {timestamp} ({user_name}) ---\n{new_note}"
        
        lead.save()
        
        # Create audit log
        from accounts.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='notes_updated',
            model_name='Lead',
            object_id=str(lead.id),
            changes={'notes': new_note},
        )
        
        # Return success response for HTMX
        # Check for HTMX request header (case-insensitive)
        is_htmx = request.headers.get('HX-Request') or request.headers.get('hx-request')
        if is_htmx:
            # Check if this is for lead detail page (target is #notes-section) or lead list (target is #notes-result)
            # HTMX sends the target in HX-Target header
            hx_target = request.headers.get('HX-Target', '') or request.headers.get('Hx-Target', '')
            
            # If target is notes-section, it's from detail page
            # If target is notes-result or empty, it's from list page
            if 'notes-section' in hx_target:
                # For lead detail page - return updated notes section
                from django.template.loader import render_to_string
                notes_html = render_to_string('leads/notes_section.html', {
                    'lead': lead,
                }, request=request)
                response = HttpResponse(notes_html)
                response['HX-Trigger'] = 'closeNotesModal'
                return response
            else:
                # For lead list page - return success message and trigger reload
                success_html = (
                    '<div class="p-4 bg-green-50 border border-green-200 rounded-lg">'
                    '<div class="flex items-center">'
                    '<svg class="w-5 h-5 text-green-600 mr-2" fill="none" stroke="currentColor" viewBox="0 0 24 24">'
                    '<path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M5 13l4 4L19 7"></path>'
                    '</svg>'
                    '<p class="text-green-800 font-medium">Note added successfully!</p>'
                    '</div>'
                    '<p class="text-green-600 text-sm mt-2">Reloading page...</p>'
                    '</div>'
                )
                response = HttpResponse(success_html)
                response['HX-Trigger'] = 'reloadPage'
                return response
        
        # Fallback: if not HTMX, return JSON (shouldn't happen but handle it)
        return JsonResponse({'success': True, 'message': 'Note added successfully'})


@login_required
def update_budget(request, pk):
    """Update lead budget via dropdown - budget is global, not project-specific"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - check associations
    has_permission = False
    if request.user.is_telecaller() or request.user.is_closing_manager():
        associations = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        )
        has_permission = associations.exists()
    else:
        has_permission = True  # Other roles can update
    
    if not has_permission:
        return JsonResponse({'success': False, 'error': 'You do not have permission to update this lead.'}, status=403)
    
    budget_value = request.POST.get('budget', '').strip()
    
    if not budget_value:
        # Clear budget
        lead.budget = None
        lead.save()
    else:
        # Budget value is already in rupees (from dropdown choices)
        try:
            from decimal import Decimal
            budget = Decimal(budget_value)
            lead.budget = budget
            lead.save()
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid budget value.'}, status=400)
    
    # Create audit log
    from accounts.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action='budget_updated',
        model_name='Lead',
        object_id=str(lead.id),
        changes={'budget': str(lead.budget) if lead.budget else 'None'},
    )
    
    messages.success(request, 'Budget updated successfully.')
    return redirect('leads:list')


@login_required
def update_configuration(request, pk):
    """Update lead configuration via dropdown"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check
    if request.user.is_telecaller() or request.user.is_closing_manager():
        if lead.assigned_to != request.user:
            return JsonResponse({'success': False, 'error': 'You do not have permission to update this lead.'}, status=403)
    
    # Get configuration IDs (multiple selection)
    configuration_ids = request.POST.getlist('configuration_ids[]') or request.POST.getlist('configuration_ids') or [request.POST.get('configuration', '')]
    configuration_ids = [cid for cid in configuration_ids if cid]  # Remove empty strings
    
    if not configuration_ids or (len(configuration_ids) == 1 and configuration_ids[0] == 'open_budget'):
        # Clear configurations
        lead.configurations.clear()
        if len(configuration_ids) == 1 and configuration_ids[0] == 'open_budget':
            # Also clear budget for open budget
            lead.budget = None
        lead.save()
    else:
        try:
            configs = GlobalConfiguration.objects.filter(id__in=configuration_ids, is_active=True)
            lead.configurations.set(configs)
            lead.save()
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Invalid configuration: {str(e)}'}, status=400)
    
    # Create audit log
    from accounts.models import AuditLog
    AuditLog.objects.create(
        user=request.user,
        action='configuration_updated',
        model_name='Lead',
        object_id=str(lead.id),
        changes={'configuration': str(lead.configuration) if lead.configuration else 'None'},
    )
    
    messages.success(request, 'Configuration updated successfully.')
    return redirect('leads:list')


@login_required
def update_lead_data(request, pk):
    """Update lead data - Only Closing Managers, Sourcing Managers, Site Heads, and Super Admins"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check - Closing Managers, Sourcing Managers, Site Heads, Mandate Owners, and Super Admins can update lead data
    if not (request.user.is_closing_manager() or request.user.is_sourcing_manager() or 
            request.user.is_site_head() or request.user.is_super_admin() or request.user.is_mandate_owner()):
        return JsonResponse({'success': False, 'error': 'You do not have permission to update lead data.'}, status=403)
    
    try:
        # Get form data - only use fields that exist in Lead model
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        age = request.POST.get('age', '').strip()
        gender = request.POST.get('gender', '').strip()
        locality = request.POST.get('locality', '').strip()
        current_residence = request.POST.get('current_residence', '').strip()
        budget = request.POST.get('budget', '').strip()
        purpose = request.POST.get('purpose', '').strip()
        visit_type = request.POST.get('visit_type', '').strip()
        visit_source = request.POST.get('visit_source', '').strip()
        how_did_you_hear = request.POST.get('how_did_you_hear', '').strip()
        occupation = request.POST.get('occupation', '').strip()
        company_name = request.POST.get('company_name', '').strip()
        designation = request.POST.get('designation', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        # Validate required fields
        if not name:
            return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)
        if not phone:
            return JsonResponse({'success': False, 'error': 'Phone is required.'}, status=400)
        
        # Update lead fields - only update fields that exist in the model
        lead.name = name
        lead.phone = phone
        if email:
            lead.email = email
        if age:
            lead.age = int(age) if age else lead.age
        if gender:
            lead.gender = gender
        if locality:
            lead.locality = locality
        if current_residence:
            lead.current_residence = current_residence
        if budget:
            lead.budget = float(budget) if budget else lead.budget
        if purpose:
            lead.purpose = purpose
        if visit_type:
            lead.visit_type = visit_type
        if visit_source:
            lead.visit_source = visit_source
        if how_did_you_hear:
            lead.how_did_you_hear = how_did_you_hear
        if occupation:
            lead.occupation = occupation
        if company_name:
            lead.company_name = company_name
        if designation:
            lead.designation = designation
        if notes:
            lead.notes = notes
        
        lead.save()
        
        # Create audit log
        from accounts.models import AuditLog
        AuditLog.objects.create(
            user=request.user,
            action='lead_data_updated',
            model_name='Lead',
            object_id=str(lead.id),
            changes={'lead_name': lead.name, 'message': 'Lead data updated via detail page'},
        )
        
        # Prepare updated data for response
        updated_data = {
            'name': lead.name,
            'phone': lead.phone,
            'email': lead.email or '-',
            'alternate_phone': lead.alternate_phone or '-',
            'whatsapp_number': lead.whatsapp_number or '-',
            'address': lead.address or '-',
            'budget_display': '',
            'lead_source_display': lead.get_visit_source_display() or '-',
            'occupation': lead.occupation or '-',
            'company': lead.company_name or '-',
            'annual_income': lead.annual_income or '-',
            'work_address': lead.work_address or '-',
        }
        
        # Format budget for display
        if lead.budget:
            if lead.budget >= 10000000:
                budget_cr = lead.budget // 10000000
                updated_data['budget_display'] = f'₹{budget_cr} Cr'
            else:
                budget_l = lead.budget // 100000
                updated_data['budget_display'] = f'₹{budget_l} L'
        else:
            updated_data['budget_display'] = '-'
        
        return JsonResponse({
            'success': True,
            'message': 'Lead data updated successfully!',
            'updated_data': updated_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Error updating lead data: {str(e)}'
        }, status=500)


@login_required
def track_call_click(request, pk):
    """Track when user clicks call button (for metrics) - creates a CallLog entry"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method.'}, status=405)
    
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    # Permission check
    has_permission = False
    if request.user.is_telecaller() or request.user.is_closing_manager():
        has_permission = lead.project_associations.filter(
            assigned_to=request.user,
            is_archived=False
        ).exists()
    else:
        has_permission = True
    
    if not has_permission:
        return JsonResponse({'success': False, 'error': 'You do not have permission to track calls for this lead.'}, status=403)
    
    # Create a CallLog entry for tracking call button clicks
    from django.utils import timezone
    CallLog.objects.create(
        lead=lead,
        user=request.user,
        call_date=timezone.now(),
        outcome='call_initiated',  # Special outcome for call button clicks
        notes='Call button clicked',
    )
    
    return JsonResponse({'success': True})


@login_required
def whatsapp(request, pk):
    """Generate WhatsApp link with template"""
    lead = get_object_or_404(Lead, pk=pk, is_archived=False)
    
    template_key = request.GET.get('template', 'intro')
    templates = get_whatsapp_templates()
    
    message = templates.get(template_key, templates['intro'])
    
    # Get primary association for project details
    primary_association = lead.project_associations.filter(is_archived=False).first()
    project = primary_association.project if primary_association else None
    
    # Replace placeholders based on template type
    if template_key == 'booking_confirmation':
        # Try to get booking ID if exists
        try:
            booking = lead.bookings.first()  # Get first booking if multiple exist
            if booking:
                message = message.format(booking_id=booking.id, name=lead.name, project_name=booking.project.name if booking.project else 'N/A')
            else:
                message = message.format(booking_id='N/A', name=lead.name, project_name=project.name if project else 'N/A')
        except:
            message = message.format(booking_id='N/A', name=lead.name, project_name=project.name if project else 'N/A')
    elif template_key == 'pretag' or template_key == 'at_site':
        # Pretag and At Site templates for Sourcing Manager
        from datetime import datetime
        visit_date = primary_association.visit_scheduled_date.strftime('%d/%m/%Y') if primary_association and primary_association.visit_scheduled_date else datetime.now().strftime('%d/%m/%Y')
        visit_time = primary_association.time_frame if primary_association and primary_association.time_frame else 'N/A'
        
        configs = ', '.join([c.display_name for c in lead.configurations.all()]) if lead.configurations.exists() else '—'
        budget_display = f'₹{lead.budget/100000:.1f}L' if lead.budget else '—'
        
        cp_name = lead.channel_partner.cp_name if lead.channel_partner else (lead.cp_name or '—')
        cp_firm = lead.channel_partner.firm_name if lead.channel_partner else (lead.cp_firm_name or '—')
        
        sm_name = lead.created_by.username if lead.created_by else '—'
        
        message = message.format(
            sm_name=sm_name,
            cp_firm=cp_firm,
            project_name=project.name if project else '—',
            visit_date=visit_date,
            visit_time=visit_time,
            client_name=lead.name,
            client_phone=lead.phone if primary_association and primary_association.phone_verified else '—',
            requirement=configs,
            budget=budget_display,
            residing_location=lead.locality or '—',
            salaried='Yes' if lead.occupation == 'service' else 'No',
            annual_income='N/D',
            cp_name=cp_name
        )
    elif template_key == 'closing_manager':
        # Closing Manager template
        from datetime import datetime
        visited_on = primary_association.last_visit_date.strftime('%d/%m/%Y') if primary_association and primary_association.last_visit_date else datetime.now().strftime('%d/%m/%Y')
        
        cm_name = primary_association.assigned_to.username if primary_association and primary_association.assigned_to else '—'
        cp_firm = lead.channel_partner.firm_name if lead.channel_partner else (lead.cp_firm_name or '—')
        sm_name = lead.created_by.username if lead.created_by else '—'
        
        configs = ', '.join([c.display_name for c in lead.configurations.all()]) if lead.configurations.exists() else '—'
        budget_display = f'₹{lead.budget/100000:.1f}L – ₹{(lead.budget*1.1)/100000:.1f}L' if lead.budget else '—'
        
        # Get notes
        notes = lead.notes or '—'
        if notes and len(notes) > 200:
            notes = notes[:200] + '...'
        
        message = message.format(
            client_name=lead.name,
            cm_name=cm_name,
            cp_firm=cp_firm,
            sm_name=sm_name,
            visited_on=visited_on,
            current_residence=lead.current_residence or '—',
            residence_location=lead.locality or '—',
            work_location=lead.locality or '—',
            ethnicity='—',
            typology=configs,
            carpet_area='—',  # Would need to get from booking/unit
            ocr='20%',
            loan_requirement='Yes' if lead.budget else 'No',
            loan_eligibility='N/A',
            purpose=lead.purpose or '—',
            budget=budget_display,
            tco_offered='Offer cost sheet',
            tco_expectation=budget_display,
            remarks=notes,
            senior_intervention='No'
        )
    else:
        # Replace placeholders for other templates
        message = message.format(name=lead.name, project_name=project.name if project else 'N/A')
    
    whatsapp_link = get_whatsapp_link(lead.phone, message)
    
    return redirect(whatsapp_link)


@login_required
def lead_assign(request):
    """Assign leads to employees (Site Head only)"""
    if not request.user.is_site_head():
        messages.error(request, 'Only Site Heads can assign leads.')
        return redirect('dashboard')
    
    # Get projects for this site head
    projects = Project.objects.filter(site_head=request.user, is_active=True)
    
    if request.method == 'POST':
        try:
            project_id = request.POST.get('project')
            if not project_id:
                messages.error(request, 'Please select a project.')
                return redirect('leads:assign')
            
            project = get_object_or_404(Project, pk=project_id, site_head=request.user)
            
            # Get assignment data
            assignments = []
            for key, value in request.POST.items():
                if key.startswith('employee_') and value:
                    employee_id = key.replace('employee_', '')
                    num_leads = int(value)
                    if num_leads > 0:
                        assignments.append({
                            'employee_id': employee_id,
                            'num_leads': num_leads
                        })
            
            if not assignments:
                messages.error(request, 'Please assign at least one lead to an employee.')
                return redirect('leads:assign')
            
            # Get employees - Site Head only sees employees assigned to their projects
            if request.user.is_site_head():
                # Get employees assigned to this site head's projects
                site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
                employees = User.objects.filter(
                    Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
                    assigned_projects__in=site_head_projects,
                    is_active=True
                ).distinct()
            else:
                # Super Admin and Mandate Owner see all employees
                employees = User.objects.filter(
                    Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
                    is_active=True
                )
            
            # Use atomic transaction with row locking to prevent race conditions
            from django.db import transaction
            with transaction.atomic():
                assigned_count = 0
                
                for assignment in assignments:
                    employee = get_object_or_404(employees, pk=assignment['employee_id'])
                    num_leads = assignment['num_leads']
                    
                    # Use select_for_update to lock rows and prevent concurrent assignment
                    leads_to_assign = list(
                        Lead.objects.select_for_update(skip_locked=True)
                        .filter(
                            project=project,
                            assigned_to__isnull=True,
                            is_archived=False
                        )
                        .order_by('created_at')[:num_leads]
                    )
                    
                    for lead in leads_to_assign:
                        lead.assigned_to = employee
                        lead.assigned_by = request.user
                        lead.assigned_at = timezone.now()
                        lead.save()
                        assigned_count += 1
                        
                        # Create audit log
                        from accounts.models import AuditLog
                        AuditLog.objects.create(
                            user=request.user,
                            action='lead_assigned',
                            model_name='LeadProjectAssociation',
                            object_id=str(association.id),
                            changes={'lead_name': association.lead.name, 'project': project.name, 'assigned_to': employee.username},
                        )
            
            messages.success(request, f'Successfully assigned {assigned_count} lead(s) to employees.')
            return redirect('leads:list')
            
        except Exception as e:
            messages.error(request, f'Error assigning leads: {str(e)}')
    
    # Get projects with unassigned leads count
    projects_with_counts = []
    for project in projects:
        unassigned_count = Lead.objects.filter(
            project=project,
            assigned_to__isnull=True,
            is_archived=False
        ).count()
        projects_with_counts.append({
            'project': project,
            'unassigned_count': unassigned_count
        })
    
    # Get employees - Site Head only sees employees assigned to their projects
    if request.user.is_site_head():
        # Get employees assigned to this site head's projects
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        employees = User.objects.filter(
            Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
            assigned_projects__in=site_head_projects,
            is_active=True
        ).distinct().order_by('username')
    else:
        # Super Admin and Mandate Owner see all employees
        employees = User.objects.filter(
            Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
            is_active=True
        ).order_by('username')
    
    context = {
        'projects': projects_with_counts,
        'employees': employees,
    }
    return render(request, 'leads/assign.html', context)


def _create_column_mapper(headers):
    """
    Create an intelligent column mapper that auto-detects column names.
    Returns a function that can extract values by field name.
    """
    # Normalize headers: strip, lowercase, remove special chars
    normalized_headers = {}
    for idx, header in enumerate(headers):
        if header:
            normalized = str(header).strip().lower().replace('_', ' ').replace('-', ' ')
            normalized_headers[normalized] = idx
    
    # Define field mappings with common variations
    # Note: This is for LEADS upload only. Visits are created separately when leads actually visit.
    field_mappings = {
        'name': ['name', 'full name', 'client name', 'customer name', 'person name', 'contact name', 'lead name'],
        'phone': ['phone', 'mobile', 'contact', 'contact number', 'phone number', 'mobile number', 'cell', 'cell phone', 'whatsapp', 'whatsapp number'],
        'email': ['email', 'e mail', 'email address', 'mail', 'email id'],
        'age': ['age'],
        'gender': ['gender', 'sex'],
        'locality': ['locality', 'area', 'location', 'city', 'address'],
        'current_residence': ['current residence', 'residence', 'residence type', 'living in', 'own rent'],
        'occupation': ['occupation', 'profession', 'job', 'work'],
        'company_name': ['company name', 'company', 'organization', 'org', 'firm name'],
        'designation': ['designation', 'position', 'title', 'role', 'job title'],
        'budget': ['budget', 'price range', 'budget range', 'expected budget', 'investment amount'],
        'purpose': ['purpose', 'requirement', 'need', 'buying purpose'],
        'visit_type': ['visit type', 'visit', 'accompanied by', 'family alone'],
        'is_first_visit': ['first visit', 'is first visit', 'new visit', 'revisit'],
        'how_did_you_hear': ['how did you hear', 'source', 'referral source', 'lead source', 'marketing source'],
        'status': ['status', 'lead status', 'stage', 'current status', 'lead stage'],  # IMPORTANT: Lead Status
        'cp_firm_name': ['cp firm name', 'channel partner firm', 'cp firm', 'partner firm', 'broker firm'],
        'cp_name': ['cp name', 'channel partner name', 'cp', 'partner name', 'broker name'],
        'cp_phone': ['cp phone', 'channel partner phone', 'cp mobile', 'partner phone', 'broker phone'],
        'cp_rera_number': ['cp rera number', 'rera number', 'cp rera', 'rera id', 'rera'],
        'is_pretagged': ['is pretagged', 'pretagged', 'pretag', 'is pretag', 'pretagged lead'],
    }
    
    # Create reverse mapping: field -> column index
    field_to_index = {}
    for field, variations in field_mappings.items():
        for variation in variations:
            if variation in normalized_headers:
                field_to_index[field] = normalized_headers[variation]
                break
    
    def get_value(field_name):
        """Get value for a field by trying all variations"""
        if field_name in field_to_index:
            idx = field_to_index[field_name]
            if idx < len(headers):
                return headers[idx]
        return None
    
    return get_value, field_to_index


@login_required
def upload_analyze(request):
    """Analyze uploaded file and return headers with auto-mapping"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
        
        # Store file in session temporarily
        import uuid
        session_id = str(uuid.uuid4())
        
        # Read headers
        file_name = uploaded_file.name.lower()
        headers = []
        
        if file_name.endswith('.csv'):
            decoded_file = uploaded_file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            headers = list(reader.fieldnames or [])
            # Store file content in session
            request.session[f'upload_file_{session_id}'] = {
                'name': uploaded_file.name,
                'content': decoded_file,
                'type': 'csv'
            }
        else:
            if openpyxl is None:
                return JsonResponse({'success': False, 'error': 'openpyxl not installed'}, status=500)
            workbook = openpyxl.load_workbook(uploaded_file)
            worksheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
            # Store file in session (we'll need to save it temporarily)
            uploaded_file.seek(0)
            file_content = uploaded_file.read()
            request.session[f'upload_file_{session_id}'] = {
                'name': uploaded_file.name,
                'content': file_content,
                'type': 'excel'
            }
        
        request.session.modified = True
        
        # Auto-detect mapping
        get_value, field_map = _create_column_mapper(headers)
        
        # Convert field_map (index-based) to header-based mapping
        auto_mapping = {}
        for field, col_idx in field_map.items():
            if col_idx < len(headers):
                auto_mapping[headers[col_idx]] = field
        
        return JsonResponse({
            'success': True,
            'session_id': session_id,
            'headers': headers,
            'mapping': auto_mapping
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def upload_preview(request):
    """Preview upload with custom mapping"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'}, status=405)
    
    try:
        session_id = request.POST.get('session_id')
        mapping_json = request.POST.get('mapping')
        
        if not session_id or not mapping_json:
            return JsonResponse({'success': False, 'error': 'Missing parameters'}, status=400)
        
        # Get file from session
        file_data = request.session.get(f'upload_file_{session_id}')
        if not file_data:
            return JsonResponse({'success': False, 'error': 'File not found in session'}, status=400)
        
        # Parse mapping
        import json
        mapping = json.loads(mapping_json)
        
        # Count rows and validate
        total_rows = 0
        valid_rows = 0
        errors = 0
        error_rows = []  # Store error details
        
        if file_data['type'] == 'csv':
            io_string = io.StringIO(file_data['content'])
            reader = csv.DictReader(io_string)
            # Create reverse mapping: field -> header
            field_to_header = {v: k for k, v in mapping.items()}
            
            for row_num, row in enumerate(reader, start=2):
                total_rows += 1
                # Check required fields using mapping
                name_header = field_to_header.get('name', '')
                phone_header = field_to_header.get('phone', '')
                
                name = row.get(name_header, '').strip() if name_header else ''
                phone = row.get(phone_header, '').strip() if phone_header else ''
                
                # Clean phone number (remove spaces, handle multiple contacts)
                if phone:
                    phone = phone.split(',')[0].strip()  # Take first phone if multiple
                    phone = phone.replace(' ', '').replace('-', '').replace('/', '')
                    # Remove leading +91 or 91
                    if phone.startswith('+91'):
                        phone = phone[3:]
                    elif phone.startswith('91') and len(phone) > 10:
                        phone = phone[2:]
                
                # Phone is required, name is optional
                if phone:
                    valid_rows += 1
                else:
                    errors += 1
                    error_rows.append({
                        'row': row_num,
                        'error': 'Phone is required',
                        'data': dict(row)
                    })
        else:
            # Excel preview
            import tempfile
            import os
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(file_data['content'])
                tmp_path = tmp.name
            
            try:
                workbook = openpyxl.load_workbook(tmp_path)
                worksheet = workbook.active
                headers = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
                
                # Create reverse mapping: field -> header
                field_to_header = {v: k for k, v in mapping.items()}
                
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                    total_rows += 1
                    # Get name and phone using mapping
                    name_header = field_to_header.get('name', '')
                    phone_header = field_to_header.get('phone', '')
                    
                    name = ''
                    phone = ''
                    if name_header in headers:
                        name_idx = headers.index(name_header)
                        if name_idx < len(row):
                            name = str(row[name_idx].value).strip() if row[name_idx].value else ''
                    if phone_header in headers:
                        phone_idx = headers.index(phone_header)
                        if phone_idx < len(row):
                            phone = str(row[phone_idx].value).strip() if row[phone_idx].value else ''
                # Clean phone number (remove spaces, handle multiple contacts)
                if phone:
                    phone = phone.split(',')[0].strip()  # Take first phone if multiple
                    phone = phone.replace(' ', '').replace('-', '').replace('/', '')
                    # Remove leading +91 or 91
                    if phone.startswith('+91'):
                        phone = phone[3:]
                    elif phone.startswith('91') and len(phone) > 10:
                        phone = phone[2:]
                
                # Phone is required, name is optional
                if phone:
                    valid_rows += 1
                else:
                    errors += 1
                    error_rows.append({
                        'row': row_num,
                        'error': 'Phone is required',
                        'data': {headers[i]: str(row[i].value) if i < len(row) and row[i].value else '' for i in range(len(headers))}
                    })
            finally:
                os.unlink(tmp_path)
        
        return JsonResponse({
            'success': True,
            'total_rows': total_rows,
            'valid_rows': valid_rows,
            'errors': errors,
            'error_rows': error_rows[:50]  # Limit to first 50 errors for preview
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def lead_upload(request):
    """Upload leads via Excel/CSV (Admin, Mandate Owner, and Site Head only) with auto-mapping and manual mapping support"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        messages.error(request, 'Only Admins, Mandate Owners, and Site Heads can upload leads.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        try:
            project_id = request.POST.get('project')
            if not project_id:
                messages.error(request, 'Please select a project.')
                return redirect('leads:upload')
            
            project = get_object_or_404(Project, pk=project_id, is_active=True)
            
            # Permission check for Site Head
            if request.user.is_site_head() and project.site_head != request.user:
                messages.error(request, 'You can only upload leads for your assigned projects.')
                return redirect('leads:upload')
            
            # Get file
            uploaded_file = request.FILES.get('file')
            if not uploaded_file:
                messages.error(request, 'Please select a file to upload.')
                return redirect('leads:upload')
            
            # Check file extension
            file_name = uploaded_file.name.lower()
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls') or file_name.endswith('.csv')):
                messages.error(request, 'Please upload an Excel (.xlsx, .xls) or CSV file.')
                return redirect('leads:upload')
            
            # Check if manual mapping is provided
            manual_mapping_json = request.POST.get('mapping')
            session_id = request.POST.get('session_id')
            
            # Parse manual mapping if provided
            manual_mapping = {}
            if manual_mapping_json:
                import json
                try:
                    manual_mapping = json.loads(manual_mapping_json)
                except:
                    pass
            
            # Process file
            leads_created = 0
            errors = []
            error_rows = []  # Store failed rows for CSV download
            mapping_info = []
            field_map = {}
            headers = []
            
            if file_name.endswith('.csv'):
                # Process CSV
                if session_id and request.session.get(f'upload_file_{session_id}'):
                    file_data = request.session.get(f'upload_file_{session_id}')
                    decoded_file = file_data['content']
                    # Clean up session
                    del request.session[f'upload_file_{session_id}']
                else:
                    decoded_file = uploaded_file.read().decode('utf-8')
                
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                
                # Get headers from CSV
                csv_headers = reader.fieldnames or []
                headers = csv_headers
                
                # Use manual mapping if provided, otherwise auto-detect
                if manual_mapping:
                    # Manual mapping: header -> field
                    header_to_field = manual_mapping
                else:
                    get_value, field_map = _create_column_mapper(csv_headers)
                    # Convert to header->field format
                    header_to_field = {}
                    for field, col_idx in field_map.items():
                        if col_idx < len(csv_headers):
                            header_to_field[csv_headers[col_idx]] = field
                
                # Store mapping info for user feedback
                if field_map:
                    mapping_info.append(f"Detected columns: {', '.join([f'{k} → {csv_headers[field_map[k]]}' for k in ['name', 'phone'] if k in field_map])}")
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        # Use manual or auto-mapping to get values
                        def get_row_value(field_name):
                            if manual_mapping:
                                # Find header mapped to this field
                                header = None
                                for h, f in header_to_field.items():
                                    if f == field_name:
                                        header = h
                                        break
                                if header:
                                    return str(row.get(header, '')).strip() if row.get(header) else ''
                            else:
                                col_idx = field_map.get(field_name)
                                if col_idx is not None and col_idx < len(csv_headers):
                                    col_name = csv_headers[col_idx]
                                    return str(row.get(col_name, '')).strip() if row.get(col_name) else ''
                            return ''
                        
                        # Get phone (required) and name (optional)
                        phone = get_row_value('phone')
                        name = get_row_value('name')
                        
                        # Normalize phone number
                        from .utils import normalize_phone
                        if phone:
                            phone = phone.split(',')[0].strip()  # Take first phone if multiple
                            phone = normalize_phone(phone)
                        
                        # Skip row if phone is not available (don't show as error)
                        if not phone:
                            continue
                        
                        # If name is not available but phone is, generate a default name
                        if not name:
                            name = f"Lead-{phone[-4:]}"  # Use last 4 digits of phone as name
                        
                        # Get or create lead by phone (deduplication)
                        lead, lead_created = Lead.objects.get_or_create(
                            phone=phone,
                            defaults={
                                'name': name,
                                'email': get_row_value('email') or '',
                                'age': int(get_row_value('age')) if get_row_value('age') and get_row_value('age').isdigit() else None,
                                'gender': get_row_value('gender') or '',
                                'locality': get_row_value('locality') or '',
                                'current_residence': get_row_value('current_residence') or '',
                                'occupation': get_row_value('occupation') or '',
                                'company_name': get_row_value('company_name') or '',
                                'designation': get_row_value('designation') or '',
                                'created_by': request.user,
                            }
                        )
                        
                        # Update lead if it already existed
                        if not lead_created:
                            if name and lead.name != name:
                                lead.name = name
                            lead.save()
                        
                        # Get simplified fields (only what's needed for leads upload)
                        configuration_str = get_row_value('configuration')
                        budget_str = get_row_value('budget')
                        feedback = get_row_value('feedback')
                        cp_id = get_row_value('cp_id')
                        
                        # Parse configurations - match to global configurations
                        configs_to_add = []
                        if configuration_str:
                            try:
                                # Try to match to global configurations
                                normalized_config = configuration_str.replace(' ', '').replace('-', '').upper()
                                global_config = GlobalConfiguration.objects.filter(
                                    name=normalized_config,
                                    is_active=True
                                ).first()
                                
                                if not global_config:
                                    # Try partial match
                                    for gc in GlobalConfiguration.objects.filter(is_active=True):
                                        if gc.name.upper() in normalized_config or normalized_config in gc.name.upper():
                                            global_config = gc
                                            break
                                
                                if global_config:
                                    configs_to_add.append(global_config)
                            except Exception as e:
                                pass
                        
                        # Parse budget - use flexible parsing utility
                        budget = None
                        if budget_str:
                            try:
                                from leads.utils import parse_budget
                                budget = parse_budget(budget_str)
                            except Exception as e:
                                # Log error but don't fail the upload
                                budget = None
                        
                        # Get CP data if provided
                        channel_partner = None
                        is_cp_data = request.POST.get('is_cp_data', 'no') == 'yes'
                        if is_cp_data:
                            cp_id_from_form = request.POST.get('channel_partner_id')
                            if cp_id_from_form:
                                try:
                                    from channel_partners.models import ChannelPartner
                                    channel_partner = ChannelPartner.objects.get(pk=cp_id_from_form)
                                except:
                                    pass
                        elif cp_id:
                            # Try to find CP by ID
                            try:
                                from channel_partners.models import ChannelPartner
                                channel_partner = ChannelPartner.objects.filter(cp_unique_id=cp_id).first()
                            except:
                                pass
                        
                        # Get Lead Status
                        status_str = get_row_value('status')
                        # Map feedback to status if status not provided
                        if not status_str and feedback:
                            feedback_lower = feedback.lower()
                            if 'interested' in feedback_lower or 'intrested' in feedback_lower:
                                status_str = 'hot'
                            elif 'not interested' in feedback_lower or 'not intrested' in feedback_lower:
                                status_str = 'lost'
                            elif 'call back' in feedback_lower or 'callback' in feedback_lower:
                                status_str = 'contacted'
                            elif 'busy' in feedback_lower or 'not answering' in feedback_lower:
                                status_str = 'contacted'
                            elif 'already booked' in feedback_lower:
                                status_str = 'lost'
                            else:
                                status_str = 'contacted'
                        
                        # Validate status against choices
                        valid_statuses = [choice[0] for choice in LeadProjectAssociation.LEAD_STATUS_CHOICES]
                        if status_str:
                            status_str = status_str.lower().strip()
                            status_matched = None
                            for valid_status in valid_statuses:
                                if status_str == valid_status or status_str.replace('_', ' ') == valid_status.replace('_', ' '):
                                    status_matched = valid_status
                                    break
                            if not status_matched:
                                for valid_status, display_name in LeadProjectAssociation.LEAD_STATUS_CHOICES:
                                    display_lower = display_name.lower()
                                    if status_str == display_lower or status_str in display_lower or display_lower in status_str:
                                        status_matched = valid_status
                                        break
                            status = status_matched if status_matched else 'new'
                        else:
                            status = 'new'
                        
                        # Handle pretagged leads (if CP data)
                        is_pretagged = is_cp_data and channel_partner is not None
                        
                        # Store feedback in notes
                        notes = ''
                        if feedback:
                            notes = f"Feedback: {feedback}"
                        if budget_str and budget is None:
                            # Budget is "Open Budget" or "Low Budget"
                            if notes:
                                notes += f"\nBudget: {budget_str}"
                            else:
                                notes = f"Budget: {budget_str}"
                        
                        # Add configurations to lead
                        if configs_to_add:
                            lead.configurations.set(configs_to_add)
                        
                        # Update budget if provided
                        if budget:
                            lead.budget = budget
                            lead.save()
                        
                        # Update channel partner if provided
                        if channel_partner:
                            lead.channel_partner = channel_partner
                            lead.save()
                        
                        # Create or get association for this project
                        association, assoc_created = LeadProjectAssociation.objects.get_or_create(
                            lead=lead,
                            project=project,
                            defaults={
                                'status': status,
                                'is_pretagged': is_pretagged,
                                'pretag_status': 'pending_verification' if is_pretagged else '',
                                'phone_verified': False,
                                'notes': notes,
                                'created_by': request.user,
                            }
                        )
                        
                        # Update association if it already existed
                        if not assoc_created:
                            association.status = status
                            if is_pretagged:
                                association.is_pretagged = True
                                association.pretag_status = 'pending_verification'
                            if notes:
                                association.notes = notes
                            association.save()
                        
                        leads_created += 1
                    except Exception as e:
                        error_msg = f"Row {row_num}: {str(e)}"
                        errors.append(error_msg)
                        error_rows.append({
                            'row': row_num,
                            'error': error_msg,
                            'data': {headers[i]: str(row[i].value) if i < len(row) and row[i].value else '' for i in range(len(headers))}
                        })
            else:
                # Process Excel
                if openpyxl is None:
                    messages.error(request, 'openpyxl is not installed. Please install it: pip install openpyxl')
                    return redirect('leads:upload')
                
                # Get file from session if available
                if session_id and request.session.get(f'upload_file_{session_id}'):
                    file_data = request.session.get(f'upload_file_{session_id}')
                    import tempfile
                    import os
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        tmp.write(file_data['content'])
                        tmp_path = tmp.name
                    try:
                        workbook = openpyxl.load_workbook(tmp_path)
                    finally:
                        os.unlink(tmp_path)
                    # Clean up session
                    del request.session[f'upload_file_{session_id}']
                else:
                    workbook = openpyxl.load_workbook(uploaded_file)
                
                worksheet = workbook.active
                
                # Get header row
                headers = [str(cell.value).strip() if cell.value else '' for cell in worksheet[1]]
                
                # Use manual mapping if provided, otherwise auto-detect
                if manual_mapping:
                    header_to_field = manual_mapping
                else:
                    get_value, field_map = _create_column_mapper(headers)
                    # Convert to header->field format
                    header_to_field = {}
                    for field, col_idx in field_map.items():
                        if col_idx < len(headers):
                            header_to_field[headers[col_idx]] = field
                
                # Store mapping info for user feedback
                if field_map:
                    mapping_info.append(f"Detected columns: {', '.join([f'{k} → {headers[field_map[k]]}' for k in ['name', 'phone'] if k in field_map])}")
                
                for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                    try:
                        # Use manual or auto-mapping to get values
                        def get_row_value(field_name):
                            if manual_mapping:
                                # Find header mapped to this field
                                header = None
                                for h, f in header_to_field.items():
                                    if f == field_name:
                                        header = h
                                        break
                                if header and header in headers:
                                    col_idx = headers.index(header)
                                    if col_idx < len(row):
                                        cell = row[col_idx]
                                        value = cell.value
                                        if value is not None:
                                            # Handle Excel numeric phone numbers (convert float to int, then to string)
                                            if isinstance(value, (int, float)) and field_name == 'phone':
                                                return str(int(value))
                                            return str(value).strip()
                            else:
                                col_idx = field_map.get(field_name)
                                if col_idx is not None and col_idx < len(row):
                                    cell = row[col_idx]
                                    value = cell.value
                                    if value is not None:
                                        # Handle Excel numeric phone numbers (convert float to int, then to string)
                                        if isinstance(value, (int, float)) and field_name == 'phone':
                                            return str(int(value))
                                        return str(value).strip()
                            return ''
                        
                        # Get phone (required) and name (optional)
                        phone = get_row_value('phone')
                        name = get_row_value('name')
                        
                        # Clean phone number (handle multiple contacts, remove spaces)
                        if phone:
                            phone = phone.split(',')[0].strip()  # Take first phone if multiple
                            phone = phone.replace(' ', '').replace('-', '').replace('/', '').replace('.', '')
                            # Remove leading +91 or 91
                            if phone.startswith('+91'):
                                phone = phone[3:]
                            elif phone.startswith('91') and len(phone) > 10:
                                phone = phone[2:]
                        
                        # Skip row if phone is not available (don't show as error)
                        if not phone:
                            continue
                        
                        # If name is not available but phone is, generate a default name
                        if not name:
                            name = f"Lead-{phone[-4:]}"  # Use last 4 digits of phone as name
                        
                        # Normalize phone number
                        from .utils import normalize_phone
                        if phone:
                            phone = normalize_phone(phone)
                        
                        # Get or create lead by phone (deduplication)
                        lead, lead_created = Lead.objects.get_or_create(
                            phone=phone,
                            defaults={
                                'name': name,
                                'email': get_row_value('email') or '',
                                'age': int(get_row_value('age')) if get_row_value('age') and get_row_value('age').isdigit() else None,
                                'gender': get_row_value('gender') or '',
                                'locality': get_row_value('locality') or '',
                                'current_residence': get_row_value('current_residence') or '',
                                'occupation': get_row_value('occupation') or '',
                                'company_name': get_row_value('company_name') or '',
                                'designation': get_row_value('designation') or '',
                                'created_by': request.user,
                            }
                        )
                        
                        # Update lead if it already existed
                        if not lead_created:
                            if name and lead.name != name:
                                lead.name = name
                            lead.save()
                        
                        # Get simplified fields (only what's needed for leads upload)
                        configuration_str = get_row_value('configuration')
                        budget_str = get_row_value('budget')
                        feedback = get_row_value('feedback')
                        cp_id = get_row_value('cp_id')
                        
                        # Parse configurations - match to global configurations
                        configs_to_add = []
                        if configuration_str:
                            try:
                                # Try to match to global configurations
                                normalized_config = configuration_str.replace(' ', '').replace('-', '').upper()
                                global_config = GlobalConfiguration.objects.filter(
                                    name=normalized_config,
                                    is_active=True
                                ).first()
                                
                                if not global_config:
                                    # Try partial match
                                    for gc in GlobalConfiguration.objects.filter(is_active=True):
                                        if gc.name.upper() in normalized_config or normalized_config in gc.name.upper():
                                            global_config = gc
                                            break
                                
                                if global_config:
                                    configs_to_add.append(global_config)
                            except Exception as e:
                                pass
                        
                        # Parse budget - use flexible parsing utility
                        budget = None
                        if budget_str:
                            try:
                                from leads.utils import parse_budget
                                budget = parse_budget(budget_str)
                            except Exception as e:
                                # Log error but don't fail the upload
                                budget = None
                        
                        # Get CP data if provided
                        channel_partner = None
                        is_cp_data = request.POST.get('is_cp_data', 'no') == 'yes'
                        if is_cp_data:
                            cp_id_from_form = request.POST.get('channel_partner_id')
                            if cp_id_from_form:
                                try:
                                    from channel_partners.models import ChannelPartner
                                    channel_partner = ChannelPartner.objects.get(pk=cp_id_from_form)
                                except:
                                    pass
                        elif cp_id:
                            # Try to find CP by ID
                            try:
                                from channel_partners.models import ChannelPartner
                                channel_partner = ChannelPartner.objects.filter(cp_unique_id=cp_id).first()
                            except:
                                pass
                        
                        # Get Lead Status
                        status_str = get_row_value('status')
                        # Map feedback to status if status not provided
                        if not status_str and feedback:
                            feedback_lower = feedback.lower()
                            if 'interested' in feedback_lower or 'intrested' in feedback_lower:
                                status_str = 'hot'
                            elif 'not interested' in feedback_lower or 'not intrested' in feedback_lower:
                                status_str = 'lost'
                            elif 'call back' in feedback_lower or 'callback' in feedback_lower:
                                status_str = 'contacted'
                            elif 'busy' in feedback_lower or 'not answering' in feedback_lower:
                                status_str = 'contacted'
                            elif 'already booked' in feedback_lower:
                                status_str = 'lost'
                            else:
                                status_str = 'contacted'
                        
                        # Validate status against choices
                        valid_statuses = [choice[0] for choice in LeadProjectAssociation.LEAD_STATUS_CHOICES]
                        if status_str:
                            status_str = status_str.lower().strip()
                            status_matched = None
                            for valid_status in valid_statuses:
                                if status_str == valid_status or status_str.replace('_', ' ') == valid_status.replace('_', ' '):
                                    status_matched = valid_status
                                    break
                            if not status_matched:
                                for valid_status, display_name in LeadProjectAssociation.LEAD_STATUS_CHOICES:
                                    display_lower = display_name.lower()
                                    if status_str == display_lower or status_str in display_lower or display_lower in status_str:
                                        status_matched = valid_status
                                        break
                            status = status_matched if status_matched else 'new'
                        else:
                            status = 'new'
                        
                        # Handle pretagged leads (if CP data)
                        is_pretagged = is_cp_data and channel_partner is not None
                        
                        # Store feedback in notes
                        notes = ''
                        if feedback:
                            notes = f"Feedback: {feedback}"
                        if budget_str and budget is None:
                            # Budget is "Open Budget" or "Low Budget"
                            if notes:
                                notes += f"\nBudget: {budget_str}"
                            else:
                                notes = f"Budget: {budget_str}"
                        
                        # Add configurations to lead
                        if configs_to_add:
                            lead.configurations.set(configs_to_add)
                        
                        # Update budget if provided
                        if budget:
                            lead.budget = budget
                            lead.save()
                        
                        # Update channel partner if provided
                        if channel_partner:
                            lead.channel_partner = channel_partner
                            lead.save()
                        
                        # Create or get association for this project
                        association, assoc_created = LeadProjectAssociation.objects.get_or_create(
                            lead=lead,
                            project=project,
                            defaults={
                                'status': status,
                                'is_pretagged': is_pretagged,
                                'pretag_status': 'pending_verification' if is_pretagged else '',
                                'phone_verified': False,
                                'notes': notes,
                                'created_by': request.user,
                            }
                        )
                        
                        # Update association if it already existed
                        if not assoc_created:
                            association.status = status
                            if is_pretagged:
                                association.is_pretagged = True
                                association.pretag_status = 'pending_verification'
                            if notes:
                                association.notes = notes
                            association.save()
                        
                        leads_created += 1
                    except Exception as e:
                        error_msg = f"Row {row_num}: {str(e)}"
                        errors.append(error_msg)
                        error_rows.append({
                            'row': row_num,
                            'error': error_msg,
                            'data': dict(row) if 'row' in locals() else {}
                        })
            
            # Store error rows in session for CSV download
            error_session_id = None
            if error_rows:
                import uuid
                error_session_id = str(uuid.uuid4())
                request.session[f'lead_upload_errors_{error_session_id}'] = {
                    'errors': error_rows,
                    'headers': headers if headers else (list(error_rows[0]['data'].keys()) if error_rows else [])
                }
                request.session.modified = True
            
            if leads_created > 0:
                success_msg = f'Successfully uploaded {leads_created} lead(s)!'
                if mapping_info:
                    success_msg += f" {' '.join(mapping_info)}"
                messages.success(request, success_msg)
            
            if errors:
                # Check if errors are due to missing required columns
                missing_name_phone = sum(1 for e in errors if 'Name and Phone are required' in e)
                if missing_name_phone > 0 and 'name' not in field_map and 'phone' not in field_map:
                    messages.error(request, 
                        f'Could not auto-detect "Name" and "Phone" columns. '
                        f'Please ensure your file has columns with names like: Name/Full Name/Client Name and Phone/Mobile/Contact Number. '
                        f'Found columns: {", ".join(headers[:10])}...')
                else:
                    error_msg = f"Errors: {'; '.join(errors[:10])}"  # Show first 10 errors
                    if len(errors) > 10:
                        error_msg += f" and {len(errors) - 10} more..."
                    if error_session_id:
                        from django.utils.safestring import mark_safe
                        error_msg += mark_safe(f" <a href='/leads/upload/errors/{error_session_id}/' class='underline'>Download Error CSV</a>")
                    messages.warning(request, error_msg)
            
            return redirect('leads:list')
            
        except Exception as e:
            messages.error(request, f'Error uploading file: {str(e)}')
    
    # Get available projects
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        projects = Project.objects.filter(is_active=True)
    else:  # Site Head
        projects = Project.objects.filter(site_head=request.user, is_active=True)
    
    # Get channel partners for CP selection dropdown
    from channel_partners.models import ChannelPartner
    channel_partners = ChannelPartner.objects.filter(status='active').order_by('cp_name')
    
    context = {
        'projects': projects,
        'channel_partners': channel_partners,
    }
    return render(request, 'leads/upload.html', context)


@login_required
def lead_assign_admin(request):
    """Assign leads to employees (Super Admin, Mandate Owner, and Site Head)"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        messages.error(request, 'Only Super Admins, Mandate Owners, and Site Heads can assign leads.')
        return redirect('dashboard')
    
    # Get projects - Mandate Owner has same permissions as Super Admin
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        projects = Project.objects.filter(is_active=True)
    else:  # Site Head - strict isolation
        projects = Project.objects.filter(site_head=request.user, is_active=True)
    
    if request.method == 'POST':
        try:
            project_id = request.POST.get('project')
            if not project_id:
                messages.error(request, 'Please select a project.')
                return redirect('leads:assign_admin')
            
            project = get_object_or_404(Project, pk=project_id, is_active=True)
            
            # Permission check for Site Head
            if request.user.is_site_head() and project.site_head != request.user:
                messages.error(request, 'You can only assign leads for your assigned projects.')
                return redirect('leads:assign_admin')
            
            # Get assignment data - save daily quotas
            for key, value in request.POST.items():
                if key.startswith('quota_') and value:
                    employee_id = key.replace('quota_', '')
                    daily_quota = int(value)
                    if daily_quota > 0:
                        employee = get_object_or_404(User, pk=employee_id)
                        DailyAssignmentQuota.objects.update_or_create(
                            project=project,
                            employee=employee,
                            defaults={
                                'daily_quota': daily_quota,
                                'is_active': True,
                            }
                        )
            
            # Also do immediate assignment if requested
            immediate_assign = request.POST.get('immediate_assign', 'false') == 'true'
            if immediate_assign:
                from django.db import transaction
                # Use atomic transaction with row locking to prevent race conditions
                with transaction.atomic():
                    # Get quotas
                    quotas = DailyAssignmentQuota.objects.filter(
                        project=project,
                        is_active=True
                    )
                    
                    assigned_count = 0
                    for quota in quotas:
                        # Use select_for_update to lock rows and prevent concurrent assignment
                        # Use LeadProjectAssociation instead of Lead
                        associations_to_assign = list(
                            LeadProjectAssociation.objects.select_for_update(skip_locked=True)
                            .filter(
                                project=project,
                                assigned_to__isnull=True,
                                is_archived=False
                            )
                            .select_related('lead')
                            .order_by('created_at')[:quota.daily_quota]
                        )
                        
                        for association in associations_to_assign:
                            association.assigned_to = quota.employee
                            association.assigned_by = request.user
                            association.assigned_at = timezone.now()
                            association.save()
                            assigned_count += 1
                
                if assigned_count > 0:
                    messages.success(request, f'Immediately assigned {assigned_count} lead(s) and saved daily quotas.')
                else:
                    messages.success(request, 'Daily quotas saved successfully.')
            else:
                messages.success(request, 'Daily quotas saved successfully. Leads will be auto-assigned daily.')
            
            return redirect('leads:list')
            
        except Exception as e:
            messages.error(request, f'Error saving assignment quotas: {str(e)}')
    
    # Get projects with unassigned leads count (using LeadProjectAssociation)
    projects_with_counts = []
    for project in projects:
        unassigned_count = LeadProjectAssociation.objects.filter(
            project=project,
            assigned_to__isnull=True,
            is_archived=False
        ).count()
        projects_with_counts.append({
            'project': project,
            'unassigned_count': unassigned_count
        })
    
    # Get employees - Site Head only sees employees assigned to their projects
    # Mandate Owner has same permissions as Super Admin
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        employees = User.objects.filter(
            Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
            is_active=True
        ).order_by('username')
    else:  # Site Head - only employees assigned to their projects (strict isolation)
        site_head_projects = Project.objects.filter(site_head=request.user, is_active=True)
        employees = User.objects.filter(
            Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
            assigned_projects__in=site_head_projects,
            is_active=True
        ).distinct().order_by('username')
    
    # Get existing quotas for each project-employee combination
    # Also filter employees by project assignment
    project_quotas = {}
    project_employees = {}  # Store employees per project
    for item in projects_with_counts:
        project = item['project']
        project_quotas[str(project.id)] = {}
        # Get employees assigned to this project
        project_employees[str(project.id)] = list(
            project.assigned_telecallers.filter(
                Q(role='closing_manager') | Q(role='telecaller') | Q(role='sourcing_manager'),
                is_active=True
            ).values_list('id', flat=True)
        )
        quotas = DailyAssignmentQuota.objects.filter(
            project=project,
            is_active=True
        )
        for quota in quotas:
            project_quotas[str(project.id)][str(quota.employee.id)] = quota.daily_quota
    
    import json
    context = {
        'projects': projects_with_counts,
        'employees': employees,
        'project_quotas_json': json.dumps(project_quotas),
        'project_employees_json': json.dumps(project_employees),
    }
    return render(request, 'leads/assign_admin.html', context)


@login_required
def lead_upload_errors_csv(request, session_id):
    """Download error rows as CSV"""
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or request.user.is_site_head()):
        messages.error(request, 'You do not have permission to download error files.')
        return redirect('dashboard')
    
    error_data = request.session.get(f'lead_upload_errors_{session_id}')
    if not error_data:
        messages.error(request, 'Error data not found.')
        return redirect('leads:list')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="lead_upload_errors_{session_id}.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = error_data['headers'] + ['Error']
    writer.writerow(headers)
    
    # Write error rows
    for error_row in error_data['errors']:
        row_data = [error_row['data'].get(h, '') for h in error_data['headers']]
        row_data.append(error_row['error'])
        writer.writerow(row_data)
    
    # Clean up session
    del request.session[f'lead_upload_errors_{session_id}']
    
    return response


@login_required
def revisit_visit(request):
    """Create a revisit for an existing visit - search, pre-fill, verify OTP, mark as revisit"""
    # Allow all user types to create revisits
    if not (request.user.is_super_admin() or request.user.is_mandate_owner() or 
            request.user.is_site_head() or request.user.is_sourcing_manager() or
            request.user.is_closing_manager() or request.user.is_telecaller()):
        messages.error(request, 'You do not have permission to create revisits.')
        return redirect('dashboard')
    
    # Get projects based on user role
    if request.user.is_super_admin() or request.user.is_mandate_owner():
        projects = Project.objects.filter(is_active=True).order_by('name')
    elif request.user.is_site_head():
        projects = Project.objects.filter(site_head=request.user, is_active=True).order_by('name')
    elif request.user.is_sourcing_manager():
        projects = request.user.assigned_projects.filter(is_active=True).order_by('name')
    elif request.user.is_closing_manager() or request.user.is_telecaller():
        projects = request.user.assigned_projects.filter(is_active=True).order_by('name')
    else:
        projects = Project.objects.none()
    
    if request.method == 'POST':
        try:
            # Get the existing visit association
            association_id = request.POST.get('existing_visit_id')
            if not association_id:
                messages.error(request, 'Please select an existing visit to revisit.')
                return redirect('leads:revisit_visit')
            
            existing_association = get_object_or_404(LeadProjectAssociation, id=association_id, is_archived=False)
            
            # Get form data
            revisit_reason = request.POST.get('revisit_reason', '')
            time_frame = request.POST.get('time_frame', '')
            visit_scheduled_date_str = request.POST.get('visit_scheduled_date', '')
            visit_scheduled_time_str = request.POST.get('visit_scheduled_time', '')
            add_to_queue = request.POST.get('add_to_queue') == 'true'
            
            # Parse visit date and time
            visit_scheduled_date = None
            if visit_scheduled_date_str:
                try:
                    visit_scheduled_date = datetime.strptime(visit_scheduled_date_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Invalid visit date format.')
                    return redirect('leads:revisit_visit')
            
            visit_scheduled_time = None
            if visit_scheduled_time_str:
                try:
                    visit_scheduled_time = datetime.strptime(visit_scheduled_time_str, '%H:%M').time()
                except ValueError:
                    messages.error(request, 'Invalid visit time format.')
                    return redirect('leads:revisit_visit')
            
            # Create new association for revisit
            revisit_association = LeadProjectAssociation.objects.create(
                lead=existing_association.lead,
                project=existing_association.project,
                assigned_to=existing_association.assigned_to,
                is_revisit=True,
                revisit_count=existing_association.revisit_count + 1,
                revisit_reason=revisit_reason,
                previous_visit=existing_association,
                time_frame=time_frame,
                visit_scheduled_date=visit_scheduled_date,
                visit_scheduled_time=visit_scheduled_time,
                pretag_status='pending_verification' if not add_to_queue else 'queued',
                is_pretagged=True,
                created_by=request.user,
            )
            
            # Copy configurations from original visit
            revisit_association.configurations.set(existing_association.configurations.all())
            
            # Handle queue option
            if add_to_queue:
                # Add to queue - set queue timestamp and status
                revisit_association.queued_at = timezone.now()
                revisit_association.queued_by = request.user
                revisit_association.pretag_status = 'queued'
                revisit_association.save()
                
                messages.success(request, f'Revisit for {existing_association.lead.name} has been added to the visit queue!')
                return redirect('leads:visit_detail', association_id=revisit_association.id)
            
            # Send OTP for verification (normal flow)
            otp = generate_otp()
            hashed_otp = hash_otp(otp)
            
            # Create OTP log
            OtpLog.objects.create(
                lead=existing_association.lead,
                phone=existing_association.lead.phone,
                otp_hash=hashed_otp,
                created_by=request.user,
                purpose='revisit_verification'
            )
            
            # Send OTP via SMS (you'll need to implement SMS sending)
            try:
                from .utils import send_sms
                message = f"Your revisit OTP for {existing_association.lead.name} is: {otp}. Please verify to confirm the revisit."
                send_sms(existing_association.lead.phone, message)
                messages.success(request, f'OTP sent to {existing_association.lead.phone} for revisit verification.')
            except Exception as e:
                messages.warning(request, f'OTP generated: {otp}. (SMS sending failed: {str(e)})')
            
            return redirect('leads:verify_revisit_otp', association_id=revisit_association.id)
            
        except Exception as e:
            messages.error(request, f'Error creating revisit: {str(e)}')
            return redirect('leads:revisit_visit')
    
    context = {
        'projects': projects,
    }
    return render(request, 'leads/revisit_visit.html', context)


@login_required
def search_existing_visits(request):
    """AJAX endpoint to search existing visits for revisit"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)
    
    search_query = request.GET.get('q', '').strip()
    project_id = request.GET.get('project_id', '')
    
    if not search_query:
        return JsonResponse({'results': []})
    
    # Filter associations based on user role - show all visits (not just pretagged)
    associations = LeadProjectAssociation.objects.filter(
        is_archived=False
    ).select_related('lead', 'project', 'assigned_to')
    
    # Apply project filter if specified
    if project_id:
        associations = associations.filter(project_id=project_id)
    
    # Apply user role filtering
    if request.user.is_telecaller():
        associations = associations.filter(assigned_to=request.user)
    elif request.user.is_closing_manager():
        associations = associations.filter(assigned_to=request.user)
    elif request.user.is_sourcing_manager():
        # Sourcing managers see visits in their assigned projects
        associations = associations.filter(project__in=request.user.assigned_projects.all())
    elif request.user.is_site_head():
        # Site heads see visits in their projects
        associations = associations.filter(project__site_head=request.user)
    elif request.user.is_mandate_owner():
        # Mandate owners see visits in their projects
        associations = associations.filter(project__mandate_owner=request.user)
    # Super admins see all visits (no additional filtering needed)
    
    # Search by lead name, phone, or email
    associations = associations.filter(
        Q(lead__name__icontains=search_query) |
        Q(lead__phone__icontains=search_query) |
        Q(lead__email__icontains=search_query)
    ).distinct()
    
    results = []
    for assoc in associations[:20]:  # Limit to 20 results
        results.append({
            'id': assoc.id,
            'lead_name': assoc.lead.name,
            'lead_phone': assoc.lead.phone,
            'lead_email': assoc.lead.email,
            'project_name': assoc.project.name,
            'assigned_to': assoc.assigned_to.get_full_name() or assoc.assigned_to.username,
            'visit_date': assoc.visit_scheduled_date.strftime('%Y-%m-%d') if assoc.visit_scheduled_date else '',
            'visit_time': assoc.visit_scheduled_time.strftime('%H:%M') if assoc.visit_scheduled_time else '',
            'revisit_count': assoc.revisit_count,
            'display_text': f"{assoc.lead.name} - {assoc.project.name} - {assoc.lead.phone}"
        })
    
    return JsonResponse({'results': results})


@login_required
def verify_revisit_otp(request, association_id):
    """Verify OTP for revisit and mark as confirmed"""
    association = get_object_or_404(LeadProjectAssociation, id=association_id, is_revisit=True)
    
    if request.method == 'POST':
        otp_entered = request.POST.get('otp', '')
        
        if not otp_entered:
            messages.error(request, 'Please enter the OTP.')
            return render(request, 'leads/verify_revisit_otp.html', {'association': association})
        
        # Verify OTP
        latest_otp = OtpLog.objects.filter(
            lead=association.lead,
            phone=association.lead.phone,
            purpose='revisit_verification'
        ).order_by('-created_at').first()
        
        if latest_otp and verify_otp_hash(otp_entered, latest_otp.otp_hash):
            # OTP verified - mark revisit as confirmed
            association.pretag_status = 'verified'
            association.is_pretagged = False
            association.save()
            
            # Update original visit's revisit count
            if association.previous_visit:
                association.previous_visit.revisit_count = association.revisit_count
                association.previous_visit.save()
            
            messages.success(request, f'Revisit for {association.lead.name} has been verified and confirmed!')
            return redirect('leads:visit_detail', association_id=association.id)
        else:
            messages.error(request, 'Invalid OTP. Please try again.')
    
    context = {
        'association': association,
        'lead': association.lead,
    }
    return render(request, 'leads/verify_revisit_otp.html', context)


@login_required
def visit_detail(request, association_id):
    """Show details of a visit (original or revisit)"""
    association = get_object_or_404(LeadProjectAssociation, id=association_id, is_archived=False)
    
    # Check permissions based on user role
    if not (request.user.is_super_admin() or request.user.is_mandate_owner()):
        if request.user.is_telecaller() and association.assigned_to != request.user:
            messages.error(request, 'You can only view your assigned visits.')
            return redirect('dashboard')
        elif request.user.is_closing_manager() and association.assigned_to != request.user:
            messages.error(request, 'You can only view your assigned visits.')
            return redirect('dashboard')
        elif request.user.is_sourcing_manager():
            if not request.user.assigned_projects.filter(id=association.project.id).exists():
                messages.error(request, 'You can only view visits in your assigned projects.')
                return redirect('dashboard')
        elif request.user.is_site_head():
            if association.project.site_head != request.user:
                messages.error(request, 'You can only view visits in your projects.')
                return redirect('dashboard')
    
    # Get all revisits for this original visit
    revisits = LeadProjectAssociation.objects.filter(
        previous_visit=association,
        is_archived=False
    ).order_by('-created_at')
    
    context = {
        'association': association,
        'lead': association.lead,
        'project': association.project,
        'assigned_to': association.assigned_to,
        'revisits': revisits,
    }
    return render(request, 'leads/visit_detail.html', context)


@login_required
def resend_revisit_otp(request, association_id):
    """Resend OTP for revisit verification"""
    if not request.is_ajax():
        return JsonResponse({'error': 'Invalid request'}, status=400)
    
    association = get_object_or_404(LeadProjectAssociation, id=association_id, is_revisit=True)
    
    try:
        # Generate new OTP
        otp = generate_otp()
        hashed_otp = hash_otp(otp)
        
        # Create new OTP log
        OtpLog.objects.create(
            lead=association.lead,
            phone=association.lead.phone,
            otp_hash=hashed_otp,
            created_by=request.user,
            purpose='revisit_verification'
        )
        
        # Send OTP via SMS
        try:
            from .utils import send_sms
            message = f"Your revisit OTP for {association.lead.name} is: {otp}. Please verify to confirm the revisit."
            send_sms(association.lead.phone, message)
            return JsonResponse({'success': True, 'message': 'OTP sent successfully'})
        except Exception as e:
            # If SMS fails, return OTP in response for development
            return JsonResponse({'success': True, 'message': f'OTP generated: {otp} (SMS sending failed)'})
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
